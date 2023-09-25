"""Make Part Number Information CSV File."""

# Standard Library
# import copy
# import csv
import enum
import os
import re
# import shutil
# import subprocess
import sys
from csv import writer
from tkinter import filedialog

# 3rd partyライブラリ
import openpyxl


# Define Global Constant
NORMAL = 0
ERROR = 1


# Define Global Variable


# Define Function
def main():
    """Call Main function."""
    # Refer Global Variable

    # Define Local Constant
    ROOT_DIR = "C:/"

    # Define Local Variable
    work_dir = ""
    excel_file_typ = []
    excel_file_dir = ""
    excel_file_path = ""
    csv_file_path = ""
    part_number_list = []

    print("\n-----Start Process-----\n")

    print("Please Input Work Dir.")
    print("Example:C:/git/PythonTool/BuryPartNumber/work")
    work_dir = filedialog.askdirectory(initialdir=ROOT_DIR)
    # work_dir = input()
    # Debug用ファイルパス設定
    # work_dir = "C:/git/PythonTool/BuryPartNumber/work"
    # print("Work Dir:", work_dir)
    # if (os.path.exists(work_dir) is False):
    #     os.makedirs(work_dir)

    # Open Part Number Excel File
    print("Please Input Part Number Excel File.")
    excel_file_typ = [("Excel File", "*.xlsx")]
    excel_file_dir = work_dir
    # Debug用ファイルパス設定
    # excel_file_path = excel_file_dir + "/RG16RJ08_DID_List_230405_ISZ回答.xlsx"
    # try:
    excel_file_path = filedialog.askopenfilename(filetypes=excel_file_typ, initialdir=excel_file_dir)
    # except Exception as e:
    #     print("存在しないまたはアクセスできないディレクトリを指定しています。")
    #     print(e)
    #     sys.exit(1)

    print("Set Excel File Path.")
    print(excel_file_path)

    # Get Part Number Information
    print("Get Part Number Information From Excel File.")
    part_number_list = get_part_number(excel_file_path)

    # Output CSV File
    print("Output Part Number Information CSV File.")
    csv_file_path = work_dir + "/part_num.csv"
    output_csv(csv_file_path, part_number_list)
    print("csv_file_path\n")

    print("\n-----End Process-----\n")


def get_part_number(excel_file_path):
    """Get Part Number Information From Excel File."""
    # Refer Global Variable

    # Define Local Constant
    PATTERN_PART_NUMBER_DATA = r"[^a-fA-F\d]"
    ROW_INIT = 1                               # Default Row
    ROW_PROJECT_INFO = 1                       # Project Info Row
    ROW_PART_NUMBER = 9                        # Part Number Row
    COLUMN_INIT = 1                            # Default Column
    COLUMN_PROJECT_CATEGORY = 8                # Project Category Column

    class ExcelColumn(enum.Enum):
        """Column Of Part Number Matrix."""

        BLANK = enum.auto()                    # 1
        MODE = enum.auto()                     # 2
        CATEGORY = enum.auto()                 # 3
        ID = enum.auto()                       # 4
        PART_NUMBER_NAME = enum.auto()         # 5
        DATA_FORMAT = enum.auto()              # 6
        DATA_LENGTH = enum.auto()              # 7
        TARGET = enum.auto()                   # 8

    # Define Local Variable
    sheets = []
    sheet_count = 0
    tmp_sheet = ""
    sheet_name = ""
    row_index = ROW_INIT
    column_index = COLUMN_INIT
    target_column_list = []
    target_column_num = 0
    target_part_number_search_flag = False
    project_search_flag = False
    compiletype_search_flag = False
    project_name_list = []
    compile_type_list = []
    id = ""
    part_number_name = ""
    part_number_data = ""
    tmp_part_number_data = ""
    part_number_data_length = ""
    part_number_data_format = ""
    part_number_target = ""
    header_list = ["Project", "Compile Type"]
    part_number_name_list = []
    part_number_data_length_list = []
    part_number_data_format_list = []
    part_number_target_list = []
    part_number_data_list = []
    part_number_name_2d_list = []
    part_number_data_length_2d_list = []
    part_number_data_format_2d_list = []
    part_number_target_2d_list = []
    part_number_data_2d_list = []
    all_part_number_data_list = []
    all_part_number_data_2d_list = []

    print("\n-----Start Get Part Number-----\n")

    # Open Excel File
    wb_part_number = openpyxl.load_workbook(excel_file_path, data_only=True)
    sheets = wb_part_number.sheetnames

    for tmp_sheet in sheets:
        sheet_name = tmp_sheet
        ws_wb_part_number = wb_part_number[sheet_name]

        # Get Target Column
        row_index = ROW_PROJECT_INFO
        column_index = COLUMN_PROJECT_CATEGORY
        target_column_list = []
        target_column_num = 0
        target_part_number_search_flag = False
        project_name_list = []
        compile_type_list = []

        while (ws_wb_part_number.cell(row=row_index, column=COLUMN_PROJECT_CATEGORY).value is not None):
            if (ws_wb_part_number.cell(row=row_index, column=COLUMN_PROJECT_CATEGORY).value == "mot埋め込み対象設定"):
                while (ws_wb_part_number.cell(row=row_index, column=column_index + 1).value is not None):
                    if (ws_wb_part_number.cell(row=row_index, column=column_index + 1).value == "○"):
                        target_column_list.append(column_index + 1)
                        target_column_num += 1
                        target_part_number_search_flag = True
                    column_index += 1
            row_index += 1
        # mot埋め込み対象設定欄のエラーチェック
        if (target_part_number_search_flag is False):
            print("mot埋め込み対象設定情報（○/-）が正しく設定されていないため品番情報を取得できません。")
            print("Excelファイルのフォーマット及び記載内容が正しいか確認して下さい。")
            os.system("PAUSE")
            sys.exit(1)

        # Get Project Information
        for index in range(target_column_num):
            row_index = ROW_PROJECT_INFO
            project_search_flag = False
            compiletype_search_flag = False

            while (ws_wb_part_number.cell(row=row_index, column=COLUMN_PROJECT_CATEGORY).value is not None):
                if (ws_wb_part_number.cell(row=row_index, column=COLUMN_PROJECT_CATEGORY).value == "Project"):
                    if (ws_wb_part_number.cell(row=row_index, column=target_column_list[index]).value is not None):
                        project_name_list.append(ws_wb_part_number.cell(row=row_index, column=target_column_list[index]).value)
                        project_search_flag = True
                elif (ws_wb_part_number.cell(row=row_index, column=COLUMN_PROJECT_CATEGORY).value == "コンパイル種別"):
                    if (ws_wb_part_number.cell(row=row_index, column=target_column_list[index]).value is not None):
                        compile_type_list.append(ws_wb_part_number.cell(row=row_index, column=target_column_list[index]).value)
                        compiletype_search_flag = True
                else:
                    # Nothing
                    pass

                row_index += 1

        # Project設定欄のエラーチェック
        if (project_search_flag is False):
            print("プロジェクト情報が正しく設定されていないため品番情報を取得できません。")
            print("Excelファイルのフォーマット及び記載内容が正しいか確認して下さい。")
            os.system("PAUSE")
            sys.exit(1)
        # コンパイル種別設定欄のエラーチェック
        elif (compiletype_search_flag is False):
            print("コンパイル種別情報（ALL/P/C/R）が正しく設定されていないため品番情報を取得できません。")
            print("Excelファイルのフォーマット及び記載内容が正しいか確認して下さい。")
            os.system("PAUSE")
            sys.exit(1)
        else:
            # Nothing
            pass

        # Get Part Number Information
        for index in range(target_column_num):
            row_index = ROW_PART_NUMBER + 1
            part_number_name_list = []
            part_number_data_format_list = []
            part_number_data_length_list = []
            part_number_target_list = []
            part_number_data_list = []
            all_part_number_data_list = []

            while (ws_wb_part_number.cell(row=row_index, column=ExcelColumn.TARGET.value).value is not None):
                if (ws_wb_part_number.cell(row=row_index, column=ExcelColumn.TARGET.value).value == "○"):
                    if ((sheet_count < 1) and (index < 1)):
                        id = ws_wb_part_number.cell(row=row_index, column=ExcelColumn.ID.value).value
                        header_list.append(id)
                    part_number_name = ws_wb_part_number.cell(row=row_index, column=ExcelColumn.PART_NUMBER_NAME.value).value
                    part_number_data_format = ws_wb_part_number.cell(row=row_index, column=ExcelColumn.DATA_FORMAT.value).value
                    part_number_data_length = ws_wb_part_number.cell(row=row_index, column=ExcelColumn.DATA_LENGTH.value).value
                    part_number_target = ws_wb_part_number.cell(row=row_index, column=ExcelColumn.TARGET.value).value
                    part_number_data = ws_wb_part_number.cell(row=row_index, column=target_column_list[index]).value
                    tmp_part_number_data = re.sub("0x", "", part_number_data)
                    part_number_data = re.sub(PATTERN_PART_NUMBER_DATA, "", tmp_part_number_data)

                    part_number_name_list.append(part_number_name)
                    part_number_data_format_list.append(part_number_data_format)
                    part_number_data_length_list.append(part_number_data_length)
                    part_number_target_list.append(part_number_target)
                    part_number_data_list.append(part_number_data)

                row_index += 1

            part_number_name_2d_list.insert(index, part_number_name_list)
            part_number_data_format_2d_list.insert(index, part_number_data_format_list)
            part_number_data_length_2d_list.insert(index, part_number_data_length_list)
            part_number_target_2d_list.insert(index, part_number_target_list)
            part_number_data_2d_list.insert(index, part_number_data_list)

            all_part_number_data_list.append(project_name_list[index])
            all_part_number_data_list += compile_type_list[index]
            all_part_number_data_list += part_number_data_list
            all_part_number_data_2d_list.append(all_part_number_data_list)

        sheet_count += 1

    # CSVの見出しヘッダ情報挿入
    all_part_number_data_2d_list.insert(0, header_list)

    print("Part Number Information\n")
    print(all_part_number_data_2d_list)

    # Close Excel File
    wb_part_number.close()

    print("\n-----End Get Part Number-----\n")

    return (all_part_number_data_2d_list)


def output_csv(csv_file_path, part_number_list):
    """Output Part Number Information To CSV File."""
    # Refer Global Variable

    # Define Local Constant

    # Define Local Variable

    print("\n-----Start Output CSV-----\n")

    try:
        with open(csv_file_path, "w", newline="") as f_object:
            writer_object = writer(f_object)
            writer_object.writerows(part_number_list)
    except Exception as e:
        print("CSVファイルを新規に作成できません。")
        print("CSVファイルが存在するか確認してください。")
        print(e)
        sys.exit(1)

    print("\n-----End Output CSV-----\n")


if __name__ == "__main__":
    main()
