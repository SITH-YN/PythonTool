"""Auto Test."""

# Standard Library
import copy
# import csv
import enum
import os
# import re
import shutil
import subprocess
from tkinter import filedialog

# 3rd partyライブラリ
import openpyxl


# Define Global Constant
NORMAL = 0
ERROR = 1


# Define Global Variable
all_svn_rev_list: list[str] = []
all_test_info_list: list[str] = []
auto_test_info_list: list[str] = []
test_soft_dir_list: list[str] = []


# Define Function
def main():
    """Call Main function."""
    # Refer Global Variable
    global all_svn_rev_list
    global all_test_info_list

    # Define Local Constant

    # Define Local Variable
    smr_dir = ""
    excel_file_typ = []
    excel_file_dir = ""
    excel_file_path = ""
    svn_dir = ""
    work_dir = ""
    soft_dir = ""
    build_dir = ""
    result_dir = ""

    print("\n-----Start Auto Test-----\n")

    # Input Dir Information
    print("Please Input SMR Dir.")
    print("Example://sdnavi-data02.d-dengi.aisin-aw.co.jp/SDFolder/SDNavi/SD_ISUZU_LB19/SMR-ISUZULB19-00035")
    smr_dir = input()
    # smr_dir = "C:/git/PythonTool/AutoTest/SMR-xxxx-xxxxx"
    print(smr_dir)

    # Open Functional Test Item Excel File
    print("Please Input Functional Test Item Excel File.")
    excel_file_typ = [("Excel File", "*.xlsx")]
    excel_file_dir = smr_dir
    excel_file_path = filedialog.askopenfilename(filetypes=excel_file_typ, initialdir=excel_file_dir)
    # excel_file_path = excel_file_dir + "/M-01/FunctionalTestItem/FunctionalTestItems_SMR-xxxx-xxxxx.xlsx"
    print("Set Functional Test Item Excel File.")
    print(excel_file_path)

    print("Please Input SVN Repository Dir(Under .svn).")
    print("Example:C:/svn/L-B/ISUZU/L-B19/trunk")
    svn_dir = input()
    # svn_dir = "C:/svn/L-B/ISUZU/L-B19/trunk"
    print(svn_dir)

    print("Please Input Work Dir.")
    print("Example:C:/git/PythonTool/AutoTest")
    work_dir = input()
    # work_dir = "C:/git/PythonTool/AutoTest"
    soft_dir = work_dir + "/Soft"
    build_dir = work_dir + "/Build"
    result_dir = work_dir + "/Result"
    print(soft_dir)
    print(build_dir)
    print(result_dir)

    # Read Functional Test Item
    read_functional_test_item(excel_file_path)
    # Prepare Test Soft
    prepare_test_soft(svn_dir, soft_dir)
    # Run Test
    run_test(build_dir, result_dir)

    print("\n-----End Auto Test-----\n")


def read_functional_test_item(excel_file_path):
    """Read Functional Test Item."""
    # Refer Global Variable
    global all_svn_rev_list
    global all_test_info_list

    # Define Local Constant
    SHEET_TEST_SUMMARY = "JP)Test Items（M-13,14)"
    ROW_INIT = 8                                    # Default Row

    class ExcelColumn(enum.Enum):
        """Column Of Function Test Item Excel File."""

        TEST_NO = enum.auto()             # 1
        TEST_DESCRIPTION = enum.auto()    # 2
        TEST_CONDITION = enum.auto()      # 3
        TEST_ACTION = enum.auto()         # 4
        EXPECTED_RESULT = enum.auto()     # 5
        REMARK_TEST_CASE = enum.auto()    # 6
        TEST_RESULT = enum.auto()         # 7
        CHART_NO = enum.auto()            # 8
        REMARK_TEST_RESULT = enum.auto()  # 9
        AUTO_TEST_TARGET = enum.auto()    # 10

    # Define Local Variable
    i = 0
    row_start = ROW_INIT
    auto_test_flag = False
    test_no = ""
    tmp_test_action = ""
    test_action = []
    debug_flag = False
    debug_index = 0
    tmp_debug_title = ""
    tmp_debug_dir = ""
    tmp_debug_file = ""
    tmp_debug_line = ""
    tmp_debug_process = ""
    debug_title = ""
    debug_dir = ""
    debug_file = ""
    debug_line = ""
    debug_process = ""
    build_index = 0
    tmp_build_exe_dir = ""
    tmp_build_command = ""
    build_exe_dir = ""
    build_command = ""
    tmp_svn_rev = ""
    svn_rev = []
    svn_rev_no = ""

    print("\n-----Start Read Functional Test Item-----\n")

    # Read Functional Test Item & Generate Auto Test List
    print("Open Functional Test Item Excel File.")
    wb_functional_test_item = openpyxl.load_workbook(excel_file_path, data_only=True)
    ws_functional_test_item = wb_functional_test_item[SHEET_TEST_SUMMARY]

    # Get Functional Test Item Information
    i = 0
    while (ws_functional_test_item.cell(row=row_start, column=ExcelColumn.TEST_ACTION.value).value is not None):
        # Get Test No
        if (ws_functional_test_item.cell(row=row_start, column=ExcelColumn.TEST_NO.value).value is not None):
            test_no = str(ws_functional_test_item.cell(row=row_start, column=ExcelColumn.TEST_NO.value).value)
        else:
            # Nothing
            pass
        # Get Test Action
        if (ws_functional_test_item.cell(row=row_start, column=ExcelColumn.AUTO_TEST_TARGET.value).value == "AUTO_TEST"):
            auto_test_flag = True
            tmp_test_action = (ws_functional_test_item.cell(row=row_start, column=ExcelColumn.TEST_ACTION.value).value)
            test_action = tmp_test_action.split("\n")
            # Get Debug Information
            if ("[debug]" in test_action):
                debug_flag = True
                debug_index = test_action.index("[debug]")
                tmp_debug_title = test_action[debug_index + 1].split(":")
                tmp_debug_dir = test_action[debug_index + 2].split(":")
                tmp_debug_file = test_action[debug_index + 3].split(":")
                tmp_debug_line = test_action[debug_index + 4].split(":")
                tmp_debug_process = test_action[debug_index + 5].split(":")
                debug_title = tmp_debug_title[1]
                debug_dir = tmp_debug_dir[1]
                debug_file = tmp_debug_file[1]
                debug_line = tmp_debug_line[1]
                debug_process = tmp_debug_process[1]
            else:
                debug_flag = False
                debug_index = 0
                debug_title = ""
                debug_dir = ""
                debug_file = ""
                debug_line = ""
                debug_process = ""
            # Get Build Information
            if ("[build]" in test_action):
                build_index = test_action.index("[build]")
                tmp_build_exe_dir = test_action[build_index + 1].split(":")
                tmp_build_command = test_action[build_index + 2].split(":")
                build_exe_dir = tmp_build_exe_dir[1]
                build_command = tmp_build_command[1]
                if (".bat" in build_command):
                    # Nothing
                    pass
                else:
                    # error
                    pass
            else:
                # error
                build_index = 0
                build_exe_dir = ""
                build_command = ""
            # Get SVN Rev
            tmp_svn_rev = ws_functional_test_item.cell(row=row_start, column=ExcelColumn.REMARK_TEST_CASE.value).value
            svn_rev = tmp_svn_rev.split(".")
            svn_rev_no = svn_rev[1]
            # Generate SVN Rev List
            all_svn_rev_list.insert(i, svn_rev_no)
            # Generate Auto Test Information
            all_test_info_list.insert(i, [auto_test_flag, test_no, svn_rev_no, debug_flag, debug_title, debug_dir, debug_file, debug_line, debug_process, build_exe_dir, build_command])
        else:
            auto_test_flag = False

        i += 1
        row_start += 1

    wb_functional_test_item.close()

    print("\n-----End Read Functional Test Item-----\n")


def prepare_test_soft(svn_dir, soft_dir):
    """Prepare Test Soft."""
    # Refer Global Variable
    global all_svn_rev_list
    global all_test_info_list
    global auto_test_info_list
    global test_soft_dir_list

    # Define Local Constant
    CMD_SVN_UPDATE = "svn update "     # svn update [-r rev] PATH
    CMD_SVN_EXPORT = "svn export -r "  # svn export [-r rev] [--ignore-externals] URL Export_PATH

    # Define Local Variable
    i = 0
    cmd = ""
    svn_rev_list = []
    tmp_all_test_info_list = []
    svn_rev = ""
    build_exe_dir = ""
    build_command = ""
    str_build_command = ""
    base_soft_dir = ""
    test_soft_dir = ""
    debug_flag = False
    debug_title = ""
    debug_dir = ""
    debug_file = ""
    debug_line = ""
    debug_process = ""
    debug_file_path = ""
    backup_file_path = ""
    data_lines = []

    print("\n-----Start Prepare Test Soft-----\n")

    # Update SVN
    # svn update [-r rev] PATH
    cmd = CMD_SVN_UPDATE + svn_dir
    subprocess.run(cmd.split())
    # Export SVN
    # Delete Duplicate SVN Rev
    svn_rev_list = sorted(set(all_svn_rev_list), key=all_svn_rev_list.index)
    for svn_rev in svn_rev_list:
        base_soft_dir = soft_dir + "/base_rev" + svn_rev
        # svn export [-r rev] [--ignore-externals] URL Export_PATH
        cmd = CMD_SVN_EXPORT + svn_rev + " --force " + svn_dir + " " + base_soft_dir
        subprocess.run(cmd.split())

    # Delete Duplicate Test Action
    i = 0
    tmp_all_test_info_list = copy.deepcopy(all_test_info_list)
    for test_data in tmp_all_test_info_list:
        del test_data[1]
        if test_data not in auto_test_info_list:
            auto_test_info_list.insert(i, test_data)
        i += 1

    # Make Test Soft
    # Copy Base Soft To Test Soft Dir
    i = 0
    for test_data in auto_test_info_list:
        base_soft_dir = soft_dir + "/base_rev" + svn_rev
        svn_rev = test_data[1]
        build_exe_dir = test_data[8]
        build_command = test_data[9]
        if (".bat" in build_command):
            str_build_command = build_command.replace(".bat", "")
        else:
            str_build_command = build_command
        str_build_command = str_build_command.replace(" ", "_")
        debug_flag = test_data[2]
        if (debug_flag is True):
            debug_title = test_data[3]
            if (" " in debug_title):
                debug_title = debug_title.replace(" ", "_")
            else:
                # Nothing
                pass
            test_soft_dir = soft_dir + "/rev" + svn_rev + "_debug_" + debug_title + "_" + str_build_command
        else:
            test_soft_dir = soft_dir + "/rev" + svn_rev + "_" + str_build_command
        test_soft_dir_list.insert(i, test_soft_dir)

        # Delete Dir excpet build_exe_dir
        dirs = os.listdir(base_soft_dir)
        for dir_name in dirs:
            if (dir_name != build_exe_dir):
                shutil.rmtree(base_soft_dir + "/" + dir_name)
            else:
                # Nothing
                pass

        print("Copy Base Soft")
        print(base_soft_dir)
        print(test_soft_dir)
        shutil.copytree(base_soft_dir, test_soft_dir)

        # Insert Debug Process
        if (debug_flag is True):
            debug_dir = test_soft_dir + "/" + test_data[4]
            debug_file = test_data[5]
            debug_file_path = debug_dir + "/" + debug_file
            debug_line = int(test_data[6])
            debug_process = test_data[7] + "\n"
            # Backup Debug File
            backup_file_path = debug_file_path + ".bak"
            if (os.path.exists(backup_file_path) is True):
                os.remove(backup_file_path)
            shutil.copy(debug_file_path, backup_file_path)
            # Open Debug File
            with open(debug_file_path, encoding="shift_jis") as f:
                data_lines = f.readlines()
            # Edit Debug Process
            data_lines[debug_line - 1] = debug_process
            # Save Debug File
            with open(debug_file_path, mode="w", encoding="shift_jis") as f:
                f.writelines(data_lines)
            # Delete Backup File
            os.remove(backup_file_path)
        else:
            # Nothing
            pass

        i += 1

    print("\n-----End Prepare Test Soft-----\n")


def run_test(build_dir, result_dir):
    """Run Test."""
    # Refer Global Variable
    global auto_test_info_list
    global test_soft_dir_list

    # Define Local Constant

    # Define Local Variable
    i = 0
    test_result_dir = ""
    build_exe_dir = ""
    build_command = ""
    cmd = ""

    print("\n-----Start Run Test-----\n")

    # Make Result Dir
    if (os.path.isdir(result_dir) is True):
        shutil.rmtree(result_dir)
    else:
        # Nothing
        pass

    i = 0
    for test_data in auto_test_info_list:
        # Make Build Dir
        if (os.path.isdir(build_dir) is True):
            shutil.rmtree(build_dir)
        else:
            # Nothing
            pass
        # Copy Test Soft To Build Dir
        print("Copy Test Soft")
        print(test_soft_dir_list[i])
        print(build_dir)
        shutil.copytree(test_soft_dir_list[i], build_dir)

        svn_rev = test_data[1]
        debug_flag = test_data[2]
        debug_title = test_data[3]
        build_exe_dir = test_data[8]
        build_command = test_data[9]
        if (".bat" in build_command):
            str_build_command = build_command.replace(".bat", "")
        else:
            str_build_command = build_command
        str_build_command = str_build_command.replace(" ", "_")
        if (debug_flag is True):
            if (" " in debug_title):
                debug_title = debug_title.replace(" ", "_")
            test_result_dir = result_dir + "/rev" + svn_rev + "_debug_" + debug_title + "_" + str_build_command
        else:
            test_result_dir = result_dir + "/rev" + svn_rev + "_" + str_build_command

        target_build_dir = build_dir + "/" + build_exe_dir
        os.chdir(target_build_dir)
        cmd = target_build_dir + "/" + build_command
        subprocess.run(cmd)
        os.chdir(os.path.dirname(os.path.abspath(__file__)))

        # Compress Test Result
        # 7zipで7z形式に圧縮
        compress_src_file = target_build_dir
        compress_dest_file = compress_src_file + ".7z"
        # 指定した7z形式の圧縮ファイルが既に存在している場合は既存の圧縮ファイルを削除
        if os.path.isfile(compress_dest_file) is True:
            os.remove(compress_dest_file)
        compress_7zip(compress_src_file, compress_dest_file)
        shutil.rmtree(compress_src_file)

        # Store Test Result
        # Make Test Result Dir
        if (os.path.isdir(test_result_dir) is True):
            shutil.rmtree(test_result_dir)
        else:
            # Nothing
            pass
        # Copy Build Dir To Test Result Dir
        print("Copy Test Result")
        print(build_dir)
        print(test_result_dir)
        shutil.copytree(build_dir, test_result_dir)
        shutil.rmtree(build_dir)

        i += 1

    print("\n-----End Run Test-----\n")


def compress_7zip(src_file, dest_file):
    """7-Zip File compression."""
    # Refer Global Variable

    # Define Local Constant
    TOOLPATH_X86_7ZIP = "C:/Program Files (x86)/7-Zip/7z.exe"
    TOOLPATH_X64_7ZIP = "C:/Program Files/7-Zip/7z.exe"

    # Define Local Variable
    # cmd = ""
    # completed_process
    return_code = NORMAL
    toolpath_7zip = ""

    # ファイル圧縮処理開始
    print("\n-----Start compress file-----\n")

    # 7-zipアプリケーションの設定
    if os.path.isfile(TOOLPATH_X86_7ZIP) is True:
        toolpath_7zip = TOOLPATH_X86_7ZIP
    elif os.path.isfile(TOOLPATH_X64_7ZIP) is True:
        toolpath_7zip = TOOLPATH_X64_7ZIP
    else:
        toolpath_7zip = ""
        print("Not found 7-zip appication.")
        print("Please install 7-zip.")

    if toolpath_7zip != "":
        # 週報フォルダの圧縮
        # ↓スペースのsplitではC:/Program Files部分を認識できないため直接コマンドを指定
        print("Compressing file.")
        # cmd = CMD_DECOMPRESSION + " " + dest_file + " -o " + dest_dir + " -r " + TARGET_FILE_EXTENSION
        # completed_process = subprocess.run(cmd.split(), stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        # completed_process = subprocess.run(toolpath_7zip + " x " + dest_file + " -o" + dest_dir + " -r " + TARGET_FILE_EXTENSION, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        completed_process = subprocess.run(toolpath_7zip + " a " + dest_file + " " + src_file, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        return_code = completed_process.returncode

        if return_code != NORMAL:
            print("Error:Failed file compression.")
        else:
            print("Success:Successed file compression.")

    # 圧縮ファイル解凍処理終了
    print("\n-----End compress file-----\n")


if __name__ == "__main__":
    main()
