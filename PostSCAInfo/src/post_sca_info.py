"""Post SCA information to Excel script."""

# Standard Library
import csv
import enum
import os
import re
import shutil
import subprocess
import urllib.request
from tkinter import filedialog

# 3rd party Libarary
from bs4 import BeautifulSoup

import openpyxl


# Define Global Constant
# CodeSonarサーバのパス設定
CODESONAR_SERVER_COMMON_URL = "http://cshub2v44:7340/"
NORMAL = 0
ERROR = 1

# Define Global Variable
smr_no = ""
title = ""
member_name = ""
svn_dir = ""
excel_file_name = ""
html_file_name = ""
smr_dir = ""
auto_build_project_dir = ""
auto_build_dir = ""
dest_dir = ""
dest_file = ""
base_target_ldm_file = ""
target_ldm_file = ""
local_base_rev = 0
target_svn_rev = 0
after_svn_rev = ""
before_svn_rev = ""
loc_info = {"nochange": 0, "modify": 0, "insert": 0, "delete": 0}
after_compile_warning = 0
before_compile_warning = 0
result_compile_warning = 0
after_memory_usage = {"ROM": 0, "RAM": 0}
before_memory_usage = {"ROM": 0, "RAM": 0}
result_memory_usage = {"ROM": 0, "RAM": 0}
sca_dir = ""
codesonar_dir = ""
codesonar_project_csv_file = ""
result_codesonar_warning = 0
loc_diff_status = NORMAL
compile_warning_status = NORMAL
after_compile_warning_status = NORMAL
before_compile_warning_status = NORMAL
memory_usage_status = NORMAL
after_memory_usage_status = NORMAL
before_memory_usage_status = NORMAL
codesonar_warning_status = NORMAL
after_codesonar_warning_status = NORMAL
before_codesonar_warning_status = NORMAL


# Define Function
def main():
    """Call Main function."""
    # Refer Global Variable
    global smr_no

    # Define Local Constant

    # Define Local Variable

    # 処理開始
    print("\n-----Start post SCA info to Excel file-----\n")

    # 対象SMRの設定
    print("Please input SMR No.")
    print("Example:SMR-xxxx-xxxxx")
    smr_no = input()

    # プロジェクト情報取得
    get_project_info()

    # ユーザー入力情報設定
    set_user_info()

    # SVN情報取得
    get_svn_info()

    # LOC（3連diffレポート）差分取得
    get_loc_diff_info()

    # build log情報取得
    get_build_log_info()

    # SCA情報取得
    get_sca_info()

    # Excel転記
    post2excel()

    # 処理終了
    print("\n-----End post SCA info to Excel file-----\n")


def get_project_info():
    """Get PROJECT information."""
    # Refer Global Variable
    global smr_no
    global smr_dir
    global auto_build_project_dir
    global codesonar_project_csv_file

    # Define Local Constant
    PATTERN_PROJECT = r"^SMR-\w+"
    AUTO_BUILD_SERVER_COMMON_DIR = "//intra.aisin-aw.co.jp/FSROOT/CSF/FSCI/"
    # AUTO_BUILD_SECURITY_SERVER_COMMON_DIR = "//csf-sc02/SDNaviFolder4/FSCI/"
    COMPILE_PRODUCT_DIR = "/Incremental/"
    SD_PROJECT_COMMON_DIR = "//sdnavi-data02.d-dengi.aisin-aw.co.jp/SDFolder/SDNavi/"

    # Define Local Variable
    # res_smr_no
    smr_str = ""
    vehicle_type = ""
    # fuel_type = ""
    project_type = ""
    customer = ""
    project_name = ""
    sd_project_name = ""
    auto_build_project_name = ""
    codesonar_project_url = ""
    codesonar_project_part_url = ""
    codesonar_project_part_csv_file_path = ""

    # プロジェクト情報取得処理開始
    print("\n-----Start get project inforimation-----\n")

    res_smr_no = re.match(PATTERN_PROJECT, smr_no)
    smr_str = res_smr_no.group()
    # SMRフォルダのルートフォルダ判定（プロジェクトによりルートフォルダが異なる）
    match smr_str:
        # 電動化
        # TMC
        # T-222 TCU
        case "SMR-TMC6GHTCU":
            project_type = "電動化"
            customer = "TMC"
            project_name = "T-222 TCU"
            sd_project_name = "SD_TMC_6GH_TCU"
            auto_build_project_name = "TMC_6GH_TCU"
            codesonar_project_part_url = "project/388.html"
            codesonar_project_part_csv_file_path = "project/388.csv"

        # T-222 EOPCU
        case "SMR-TMC6GHEOPCU":
            project_type = "電動化"
            customer = "TMC"
            project_name = "T-222 EOPCU"
            sd_project_name = "SD_TMC_6GH_EOPCU"
            auto_build_project_name = "TMC_6GH_EOPCU"
            # codesonar_project_part_url = "project/.html"
            # codesonar_project_part_csv_file_path = "project/.csv"

        # L-B01/L-B09/L-B10
        case "SMR-TMCLB":
            print("Please input VEHICLE TYPE(select only CONV/1MHV).")
            print("L-B01の例:「CONV」")
            vehicle_type = input()
            # print("Please input FUEL TYPE(select only PETROL/DEISEL).")
            # print("L-B01の例:「PETROL」")
            # fuel_type = input()

            # VEHICLE TYPE:CONV
            if vehicle_type == "CONV":
                # FUEL TYPE:PETROL
                # if fuel_type == "PETROL":
                # FUEL TYPE:DEISEL
                # elif fuel_type == "DEISEL":
                # FUEL TYPE:それ以外
                # else:
                project_type = "電動化"
                customer = "TMC"
                project_name = "L-B01/L-B09/L-B10"
                sd_project_name = "SD_TMC_LB"
                auto_build_project_name = "TMC_FR-DAT"
                codesonar_project_part_url = "project/414.html"
                codesonar_project_part_csv_file_path = "project/414.csv"
            # VEHICLE TYPE:1MHV
            elif vehicle_type == "1MHV":
                # FUEL TYPE:PETROL
                # if fuel_type == "PETROL":
                # FUEL TYPE:DEISEL
                # elif fuel_type == "DEISEL":
                # FUEL TYPE:それ以外
                # else:
                project_type = "電動化"
                customer = "TMC"
                project_name = "L-B13"
                sd_project_name = "SD_TMC_LB"
                auto_build_project_name = "TMC_FR-DAT"
                codesonar_project_part_url = "project/439.html"
                codesonar_project_part_csv_file_path = "project/439.csv"
            # それ以外
            else:
                project_type = "電動化"
                customer = "TMC"
                project_name = ""
                sd_project_name = "SD_TMC_LB"
                auto_build_project_name = "TMC_FR-DAT"
                codesonar_project_part_url = ""
                codesonar_project_part_csv_file_path = ""

        # L-B18 TCU
        case "SMR-TMCLB18TCU":
            project_type = "電動化"
            customer = "TMC"
            project_name = "L-B18"
            sd_project_name = "SD_TMC_L-B18_TCU"
            # auto_build_project_name = "TMC_FR-DAT"
            # codesonar_project_part_url = "project/xxx.html"
            # codesonar_project_part_csv_file_path = "project/xxx.csv"

        # L-B18 EOPCU
        case "SMR-TMCLB18EOPCU":
            project_type = "電動化"
            customer = "TMC"
            project_name = "L-B18"
            sd_project_name = "SD_TMC_L-B18_EOPCU"
            # auto_build_project_name = "TMC_FR-DAT"
            # codesonar_project_part_url = "project/xxx.html"
            # codesonar_project_part_csv_file_path = "project/xxx.csv"

        # ISUZU
        # L-B19
        case "SMR-ISUZULB19":
            project_type = "電動化"
            customer = "ISUZU"
            project_name = "L-B19"
            sd_project_name = "SD_ISUZU_LB19"
            auto_build_project_name = "ISUZU_LB19"
            codesonar_project_part_url = "project/460.html"
            codesonar_project_part_csv_file_path = "project/460.csv"

        # G-GenII
        # VCC
        # G-D83
        case "SMR-VOLVOGGENIIPF5":
            project_type = "G-GenII"
            customer = "VCC"
            project_name = "G-D83"
            sd_project_name = "SD_VOLVO_G-GENII_PF5"
            auto_build_project_name = "VOLVO_G_GenII_PF5"
            codesonar_project_part_url = "project/447.html"
            codesonar_project_part_csv_file_path = "project/447.csv"

        # PSA
        # G-D94
        case "SMR-PSAG22UPPF52":
            project_type = "G-GenII"
            customer = "PSA"
            project_name = "G-D94"
            sd_project_name = "SD_PSA_G22_UP_PF5_2"
            auto_build_project_name = "PSA_G22_UP_PF5_2"
            codesonar_project_part_url = "project/450.html"
            codesonar_project_part_csv_file_path = "project/450.csv"

        # VW
        # G-A91
        case "SMR-VWGGENIIPF5X":
            project_type = "G-GenII"
            customer = "VW"
            project_name = "G-A91"
            sd_project_name = "SD_VW_G-GENII_PF5X"
            auto_build_project_name = "VW_PF5X"
            codesonar_project_part_url = "project/446.html"
            codesonar_project_part_csv_file_path = "project/446.csv"

        # default
        case _:
            project_type = ""
            customer = ""
            project_name = ""
            sd_project_name = ""
            auto_build_project_name = ""
            codesonar_project_part_url = ""
            codesonar_project_part_csv_file_path = ""
            print("Can't get SMR information.")
            print("Exit process.")
            subprocess.call('PAUSE', shell=True)

    smr_dir = SD_PROJECT_COMMON_DIR + sd_project_name + "/" + smr_no + "/"
    auto_build_project_dir = AUTO_BUILD_SERVER_COMMON_DIR + auto_build_project_name + COMPILE_PRODUCT_DIR
    codesonar_project_url = CODESONAR_SERVER_COMMON_URL + codesonar_project_part_url
    codesonar_project_csv_file = CODESONAR_SERVER_COMMON_URL + codesonar_project_part_csv_file_path

    print("SMR No:", smr_no)
    print("PROJECT TYPE:", project_type)
    print("CUSTOMER:", customer)
    print("PROJECT:", project_name)
    print("SD PROJECT:", sd_project_name)
    print("SMR DIR:", smr_dir)
    print("AUTO BUILD SERVER DIR:", auto_build_project_dir)
    print("CodeSonar PROJECT URL:", codesonar_project_url)

    # プロジェクト情報取得処理終了
    print("\n-----End get project inforimation-----\n")


def set_user_info():
    """Set user information."""
    # Refer Global Variable
    global smr_dir
    global title
    global member_name
    global svn_dir
    global local_base_rev
    global excel_file_name
    global html_file_name

    # Define Local Constant

    # Define Local Variable
    excel_file_typ = ("", "")
    excel_file_dir = ""
    html_file_typ = ("", "")
    html_file_dir = ""

    # ユーザー入力情報設定処理開始
    print("\n-----Start set user input inforimation-----\n")

    # 案件名設定
    print("Please input title.")
    print("Example:CAN信号追加1回目")
    title = input()

    # 担当者設定
    print("Please input member name.")
    print("Example:xxxx")
    member_name = input()

    # SVN repository path設定
    print("Please input SVN repository path(.svn dir full path).")
    print("Example:C:/svn/GH/TMC/6GH_TCU/trunk")
    svn_dir = input()

    # Local base rev設定
    print("Please input local base rev(input only number).")
    print("Example:123")

    while True:
        try:
            local_base_rev = int(input())
            break
        except ValueError:
            print("Invalid input number. Try again.")

    # Excelファイルパス設定
    print("Please input post target excel file.")
    excel_file_typ = [("excelファイル", "*.xlsx")]
    excel_file_dir = smr_dir + "M-08/"
    excel_file_name = filedialog.askopenfilename(filetypes=excel_file_typ, initialdir=excel_file_dir)
    print("Set post target excel file.")
    print(excel_file_name)

    # 3連diffレポートファイルパス設定
    print("Please input diff report html file.")
    html_file_typ = [("HTMLファイル", "*.html")]
    html_file_dir = smr_dir + "M-10/"
    html_file_name = filedialog.askopenfilename(filetypes=html_file_typ, initialdir=html_file_dir)
    print("Set diff report html file.")
    print(html_file_name)

    # ユーザー入力情報設定処理終了
    print("\n-----End set user input inforimation-----\n")


def get_svn_info():
    """Get SVN information."""
    # Refer Global Variable
    global smr_dir

    # Define Local Constant
    TMP_DIR = smr_dir + "M-08/tmp/"
    SVN_LOG_FILE = TMP_DIR + "svn_log.txt"
    TMP_SVN_LOG_FILE = TMP_DIR + "tmp_svn_log.txt"

    # Define Local Variable

    # SVN情報取得処理開始
    print("\n-----Start get SVN information-----\n")

    # tmpフォルダ作成
    if os.path.isdir(TMP_DIR) is False:
        os.makedirs(TMP_DIR)

    # SVN更新
    exec_svn_update()

    # SVNログ取得
    get_svn_log(SVN_LOG_FILE)

    # SVN revリスト生成
    generate_target_svn_rev_list(SVN_LOG_FILE, TMP_SVN_LOG_FILE)

    # SVN対象rev設定
    set_target_svn_rev()

    # tmpフォルダの削除
    shutil.rmtree(TMP_DIR)

    # SVN情報取得処理終了
    print("\n-----End get SVN information-----\n")


def exec_svn_update():
    """Execute SVN update."""
    # Refer Global Variable
    global svn_dir

    # Define Local Constant
    CMD_SVN_UPDATE = "svn update"

    # Define Local Variable
    cmd = ""
    # completed_process
    return_code = NORMAL

    # SVN更新処理開始
    print("\n-----Start SVN update-----\n")

    # 対象のSVNフォルダ自動更新
    print("Execute SVN update.")
    print(svn_dir)
    cmd = CMD_SVN_UPDATE + " " + svn_dir
    completed_process = subprocess.run(cmd.split(), stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    return_code = completed_process.returncode

    if return_code != NORMAL:
        print("Error:Failed execute SVN update.")
        subprocess.call('PAUSE', shell=True)

    # SVN更新処理終了
    print("\n-----End SVN update-----\n")


def get_svn_log(svn_log_file):
    """Get SVN log."""
    # Refer Global Variable
    global svn_dir

    # Define Local Constant
    CMD_SVN_LOG = "svn log"

    # Define Local Variable
    cmd = ""
    # completed_process
    return_code = NORMAL

    # SVNログ取得処理開始
    print("\n-----Start get SVN log-----\n")

    print("Get SVN log.")
    print(svn_dir)
    cmd = CMD_SVN_LOG + " " + svn_dir
    # completed_process = subprocess.run(cmd.split(), stdout=subprocess.PIPE, stderr=subprocess.PIPE, capture_output=True, text=True)
    completed_process = subprocess.run(cmd.split(), stdin=subprocess.PIPE, stdout=open(svn_log_file, "w", encoding="cp932"))
    return_code = completed_process.returncode

    if return_code != NORMAL:
        print("Error:Failed get SVN log.")
        subprocess.call('PAUSE', shell=True)

    # SVNログ取得処理終了
    print("\n-----End get SVN log-----\n")


def generate_target_svn_rev_list(svn_log_file, tmp_svn_log_file):
    """Generate target SVN rev list."""
    # Refer Global Variable
    global smr_no

    # Define Local Constant
    PATTERN_SVN_REV = r"^r\d+"
    PATTERN_SMR_NO = r"SMR-\w+-\d+"

    # Define Local Variable
    tmp_svn_log_data = ""
    smr_no_flg = 0
    index_svn_rev = 0
    # res_svn_rev
    # res_smr_no_list
    svn_rev_list = []
    smr_no_list = []
    target_svn_rev_list = []

    # Target SVN revリスト生成処理開始
    print("\n-----Start generate target SVN rev list-----\n")

    # rev一覧の整形（破線区切りで1revのため余計な空行削除&revとSMR Noの配列保持）
    try:
        with open(svn_log_file, "r", encoding="cp932") as logfile:
            with open(tmp_svn_log_file, "w", encoding="shift_jis") as tmp_logfile:
                for line in logfile:
                    if line.strip():
                        tmp_logfile.write(line)
    except FileNotFoundError as file_open_error:
        print(file_open_error, svn_log_file)
        subprocess.call('PAUSE', shell=True)

    # SVN revとSMR Noを昇順で配列に保持
    # 最初のSVN revはSMRと紐づかないため、SMR Noを「N/Aに設定」
    # SVN logのコメントに複数回SMR Noが出てくる可能性があるため、SMRNo:SMR-xxxx-xxxxから始まる部分のSMR Noのみを抽出
    try:
        with open(tmp_svn_log_file, "r", encoding="shift_jis") as tmp_logfile:
            while True:
                tmp_svn_log_data = tmp_logfile.readline()
                if tmp_svn_log_data == "":
                    break
                if re.match(PATTERN_SVN_REV, tmp_svn_log_data):
                    res_svn_rev = re.match(PATTERN_SVN_REV, tmp_svn_log_data)
                    svn_rev_list.insert(0, res_svn_rev.group())
                    smr_no_flg = 1
                elif smr_no_flg == 1:
                    if re.match("^SMRNo:" + PATTERN_SMR_NO, tmp_svn_log_data):
                        res_smr_no_list = re.search(PATTERN_SMR_NO, tmp_svn_log_data)
                        smr_no_list.insert(0, res_smr_no_list.group())
                        smr_no_flg = 0
            smr_no_list.insert(0, "N/A")
    except FileNotFoundError as file_open_error:
        print(file_open_error, tmp_svn_log_file)
        subprocess.call('PAUSE', shell=True)

    # ログ一覧から対象SMRのrevのみを取得
    # smr_no_listから対象SMRのindexを引っ張りrev_listから対象SMRのrevのみを抽出する
    # smr_no_list配列数分ループ
    for smr_name in smr_no_list:
        # 対象SMRと一致するか判定し一致したらsvn_rev_list[index]の中身をtarget_svn_rev_listに代入
        if smr_name == smr_no:
            target_svn_rev_list.append(svn_rev_list[index_svn_rev])
        index_svn_rev += 1

    print("Target SVN rev list:", target_svn_rev_list, "\n")

    # SVN revリスト生成処理終了
    print("\n-----End generate target SVN rev list-----\n")


def set_target_svn_rev():
    """Set target SVN rev."""
    # Refer Global Variable
    global target_svn_rev

    # Define Local Constant

    # Define Local Variable

    # SVN対象rev設定処理開始
    print("\n-----Start set target SVN rev-----\n")

    print("Please select target SVN rev(input only number).")
    print("Example:123")

    while True:
        try:
            target_svn_rev = int(input())
            break
        except ValueError:
            print("Invalid input number. Try again.")

    if str(target_svn_rev) == "":
        print("Error:Failed set target SVN rev.")
        subprocess.call('PAUSE', shell=True)

    # SVN対象rev設定処理終了
    print("\n-----End set target SVN rev-----\n")


def get_loc_diff_info():
    """Get LOC diff information."""
    # Refer Global Variable
    global html_file_name
    global smr_dir
    global target_svn_rev
    global loc_diff_status

    # Define Local Constant
    class LOC(enum.Enum):
        EXPLANATION = 0           # 0
        DIR1 = enum.auto()        # 1
        DIR2 = enum.auto()        # 2
        SUM = enum.auto()         # 3
        TEXT_BLOCK = enum.auto()  # 4

    # Define Local Variable
    loc_diff_csv_file = ""
    # bsObj
    # table
    # rows
    # write_csv_file
    # writer
    # write_csv_data
    csvRow = []
    # read_csv_file
    # read_csv_data
    tmp_header1 = []
    tmp_header2 = []
    tmp_loc_nochange = []
    tmp_loc_modify = []
    tmp_loc_insert = []
    tmp_loc_delete = []

    # LOC（3連diffレポート）差分取得処理開始
    print("\n-----Start get LOC diff data-----\n")

    # LOC（3連diffレポート）ログファイル保存フォルダ作成
    loc_diff_dir = smr_dir + "M-10/LOC_diff/" + str(target_svn_rev) + "/"
    if os.path.isdir(loc_diff_dir) is False:
        os.makedirs(loc_diff_dir)
    loc_diff_csv_file = loc_diff_dir + "LOC_diff_" + str(target_svn_rev) + ".csv"

    # HTML解析
    bsObj = BeautifulSoup(open(html_file_name, encoding="shift_jis"), "html.parser")

    # テーブルを指定（before/afterのLOC差分）
    table = bsObj.findAll("table", {"class": "bordered"})[2]
    rows = table.findAll("tr")

    # CSVファイルを開く
    with open(loc_diff_csv_file, "w", encoding="shift_jis", newline="") as write_csv_file:
        writer = csv.writer(write_csv_file)
        for write_csv_data in rows:
            csvRow = []
            for cell in write_csv_data.findAll(["td", "th"]):
                csvRow.append(cell.get_text())
            writer.writerow(csvRow)

    try:
        with open(loc_diff_csv_file, "r", encoding="shift_jis", errors="", newline="") as read_csv_file:
            read_csv_data = csv.reader(read_csv_file, delimiter=",", doublequote=True, lineterminator="\r\n", quotechar='"', skipinitialspace=True)
            # tmp_header1:説明,第１フォルダーと第２フォルダー
            # tmp_header2:行数（>フォルダー?1）,行数（フォルダー?2）,行数（合計）,テキスト ブロック
            # tmp_loc_nochange[0]～[4]:変更なし,23633,23633,47266,1502
            # tmp_loc_modify[0]～[4]:変更箇所,9875,20865,30740,1463
            # tmp_loc_insert[0]～[4]:挿入箇所,0,1687,1687,24
            # tmp_loc_delete[0]～[4]:削除箇所,2158,0,2158,16
            tmp_header1 = next(read_csv_data)
            tmp_header2 = next(read_csv_data)
            tmp_loc_nochange = next(read_csv_data)
            tmp_loc_modify = next(read_csv_data)
            tmp_loc_insert = next(read_csv_data)
            tmp_loc_delete = next(read_csv_data)

            # loc_info[4(0～3)]:LOC差分情報
            # loc_info[0]:nochange
            # loc_info[1]:modify
            # loc_info[2]:insert
            # loc_info[3]:delete
            loc_info["nochange"] = int(tmp_loc_nochange[LOC.SUM.value])
            loc_info["modify"] = int(tmp_loc_modify[LOC.SUM.value])
            loc_info["insert"] = int(tmp_loc_insert[LOC.SUM.value])
            loc_info["delete"] = int(tmp_loc_delete[LOC.SUM.value])
    except FileNotFoundError as file_open_error:
        loc_diff_status = ERROR
        print(file_open_error, loc_diff_csv_file)

    # LOC（3連diffレポート）差分取得処理終了
    print("\n-----End get LOC diff data-----\n")


def get_build_log_info():
    """Get build log information."""
    # Refer Global Variable

    # Define Local Constant

    # Define Local Variable

    # build log(after)情報取得
    get_after_build_log_info()

    # build log(before)情報取得
    get_before_build_log_info()

    # build log差分情報取得
    get_diff_build_log_info()


def get_after_build_log_info():
    """Get after build log information."""
    # Refer Global Variable
    global auto_build_project_dir
    global auto_build_dir
    global target_svn_rev
    global after_svn_rev
    global after_compile_warning
    global after_memory_usage
    global after_compile_warning_status
    global after_memory_usage_status

    # Define Local Constant

    # Define Local Variable
    after_ldm_file_list = []
    compile_warning_file = ""
    rst_file = ""

    # build log(after)情報取得処理開始
    print("\n-----Start get build log informatin(after)-----\n")

    after_svn_rev = str(target_svn_rev)
    auto_build_dir = auto_build_project_dir + after_svn_rev
    after_ldm_file_list = get_autobuild_product(after_svn_rev)

    set_copy_target_ldm_file(after_ldm_file_list)

    get_target_ldm_file(after_ldm_file_list)

    # ldm圧縮ファイルの解凍
    file_decompression()

    # compile warningファイル取得
    compile_warning_file = get_compile_warning_file()

    # compile warning差分取得
    # compile warning（*_error.txt）ファイルの差分比較を行いコンパイルワーニング増加件数を取得
    after_compile_warning = get_compile_warning(compile_warning_file)

    if compile_warning_status == NORMAL:
        after_compile_warning_status = NORMAL
    else:
        after_compile_warning_status = ERROR

    # rstファイル取得
    rst_file = get_rst_file()

    # メモリ使用量（ROM/RAM）差分取得
    # rstファイルの差分比較を行いメモリ使用量（ROM/RAM）の差分を取得
    after_memory_usage = get_memory_usage(rst_file)

    if memory_usage_status == NORMAL:
        after_memory_usage_status = NORMAL
    else:
        after_memory_usage_status = ERROR

    # build log(after)情報取得処理終了
    print("\n-----End get build log informatin(after)-----\n")


def get_before_build_log_info():
    """Get before build log information."""
    # Refer Global Variable
    global auto_build_project_dir
    global auto_build_dir
    global target_svn_rev
    global before_svn_rev
    global before_compile_warning
    global before_memory_usage
    global before_compile_warning_status
    global before_memory_usage_status

    # Define Local Constant

    # Define Local Variable
    before_ldm_file_list = []
    compile_warning_file = ""
    rst_file = ""

    # build log(before)情報取得処理開始
    print("\n-----Start get build log informatin(before)-----\n")

    # build log(before)情報取得
    # after rev-1で抽出できないパターンに注意（自動ビルドサーバに存在する直前のrevを抽出する）
    # after rev-1のフォルダが存在しなければループし-1を繰り返し見つかれば抜ける
    before_svn_rev = str(target_svn_rev - 1)
    auto_build_dir = auto_build_project_dir + before_svn_rev
    while os.path.isdir(auto_build_dir) is False:
        before_svn_rev = str(int(before_svn_rev) - 1)
        auto_build_dir = auto_build_project_dir + before_svn_rev
    before_ldm_file_list = get_autobuild_product(before_svn_rev)

    get_target_ldm_file(before_ldm_file_list)

    # ldm圧縮ファイルの解凍
    file_decompression()

    # compile warningファイル取得
    compile_warning_file = get_compile_warning_file()

    # compile warning差分取得
    # compile warning（*_error.txt）ファイルの差分比較を行いコンパイルワーニング増加件数を取得
    before_compile_warning = get_compile_warning(compile_warning_file)

    if compile_warning_status == NORMAL:
        before_compile_warning_status = NORMAL
    else:
        before_compile_warning_status = ERROR

    # rstファイル取得
    rst_file = get_rst_file()

    # メモリ使用量（ROM/RAM）差分取得
    # rstファイルの差分比較を行いメモリ使用量（ROM/RAM）の差分を取得
    before_memory_usage = get_memory_usage(rst_file)

    if memory_usage_status == NORMAL:
        before_memory_usage_status = NORMAL
    else:
        before_memory_usage_status = ERROR

    # build log(before)情報取得処理終了
    print("\n-----End get build log informatin(before)-----\n")


def get_autobuild_product(svn_rev):
    """Get autobuild product."""
    # Refer Global Variable
    global auto_build_dir
    global smr_dir
    global dest_dir

    # Define Local Constant

    # Define Local Variable
    base_7z_file = ""
    ext_7z_file = ""
    base_ldm_file = ""
    ext_ldm_file = ""
    ldm_file_list = []

    # SVN UP revのbefore/afterの自動ビルド結果コピー
    # （自動ビルドサーバ→SMRのM-09/build/revフォルダ）
    # M-09プロセスのビルドログ情報（ldm配下のログフォルダ + rstファイル）の自動格納
    # 不要なldm/dummyフォルダの削除（ldm/PRJ_FOLDERのログファイルだけあればよいため）
    print("\n-----Start get autobuild product-----\n")

    # 拡張子.ldm.7zで抽出
    if os.path.isdir(auto_build_dir) is True:
        dest_dir = smr_dir + "M-09/build/" + svn_rev + "/"
        if os.path.isdir(dest_dir) is True:
            shutil.rmtree(dest_dir)
        os.makedirs(dest_dir)
        for file in os.listdir(auto_build_dir):
            base_7z_file, ext_7z_file = os.path.splitext(file)
            if ext_7z_file == ".7z":
                base_ldm_file, ext_ldm_file = os.path.splitext(base_7z_file)
                if ext_ldm_file == ".ldm":
                    ldm_file_list.append(file)
        if len(ldm_file_list) != 0:
            print("Ldm file list:", ldm_file_list)
        else:
            print("Failed get ldm file")
            print("Please check", auto_build_dir)

    print("\n-----End get autobuild product-----\n")

    return(ldm_file_list)


def set_copy_target_ldm_file(ldm_file_list):
    """Set copy target ldm file."""
    # Refer Global Variable
    global target_ldm_file

    # Define Local Constant

    # Define Local Variable

    # コピー対象ldmファイルの設定処理開始
    print("\n-----Start set copy target ldm file-----\n")

    if len(ldm_file_list) != 0:
        print("Please select target ldm file.")
        target_ldm_file = input()
    else:
        print("Failed get Target ldm file")
        print("Please check selected target ldm file name.")

    # コピー対象ldmファイルの設定処理終了
    print("\n-----End set copy target ldm file-----\n")


def get_target_ldm_file(ldm_file_list):
    """Get target ldm file."""
    # Refer Global Variable
    global target_ldm_file
    global base_target_ldm_file
    global auto_build_dir
    global dest_dir
    global dest_file

    # Define Local Constant

    # Define Local Variable
    base_7z_file = ""
    ext_7z_file = ""
    ext_target_ldm_file = ""
    src_file = ""

    # 対象ldmファイルの取得処理開始
    print("\n-----Start get target ldm file-----\n")

    if len(ldm_file_list) != 0:
        base_7z_file, ext_7z_file = os.path.splitext(target_ldm_file)
        if ext_7z_file == ".7z":
            base_target_ldm_file, ext_target_ldm_file = os.path.splitext(base_7z_file)
        src_file = auto_build_dir + "/" + target_ldm_file
        dest_file = dest_dir + target_ldm_file
        print("Copy target ldm file:", target_ldm_file)
        shutil.copy2(src_file, dest_file)

    # 対象ldmファイルの取得処理終了
    print("\n-----End get target ldm file-----\n")


def file_decompression():
    """File decompression."""
    # Refer Global Variable
    global dest_dir
    global dest_file
    global base_target_ldm_file

    # Define Local Constant
    TOOLPATH_X86_7ZIP = "C:/Program Files (x86)/7-Zip/7z.exe"
    TOOLPATH_X64_7ZIP = "C:/Program Files/7-Zip/7z.exe"
    TARGET_FILE_EXTENSION = "*.err *.log *.rst *.txt"

    # Define Local Variable
    # cmd = ""
    # completed_process
    return_code = NORMAL
    toolpath_7zip = ""

    # 圧縮ファイル解凍処理開始
    print("\n-----Start decompression compressed file-----\n")

    # 7-zipアプリケーションの設定
    if os.path.isfile(TOOLPATH_X86_7ZIP) is True:
        toolpath_7zip = TOOLPATH_X86_7ZIP
    elif os.path.isfile(TOOLPATH_X64_7ZIP) is True:
        toolpath_7zip = TOOLPATH_X64_7ZIP
    else:
        toolpath_7zip = ""
        print("Not found 7-zip appication.")
        print("Please install 7-zip.")

    if toolpath_7zip != "":
        # ldm圧縮ファイルの解凍
        # ldm配下の各種ログファイル + rstファイルのみ抽出
        # ↓スペースのsplitではC:/Program Files部分を認識できないため直接コマンドを指定
        print("Decompression compressed file.")
        # cmd = CMD_DECOMPRESSION + " " + dest_file + " -o " + dest_dir + " -r " + TARGET_FILE_EXTENSION
        # completed_process = subprocess.run(cmd.split(), stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        completed_process = subprocess.run(toolpath_7zip + " x " + dest_file + " -o" + dest_dir + " -r " + TARGET_FILE_EXTENSION, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        return_code = completed_process.returncode

        if return_code != NORMAL:
            print("Error:Failed file decompression.")

        # 不要なldm/dummyフォルダの削除（ldm/PRJ_FOLDERのログファイルだけあればよいため）
        shutil.rmtree(dest_dir + base_target_ldm_file + "/ldm/dummy/")

    # 圧縮ファイル解凍処理終了
    print("\n-----End decompression compressed file-----\n")


def get_compile_warning_file():
    """Get compile warning file."""
    # Refer Global Variable
    global dest_dir
    global base_target_ldm_file
    global compile_warning_status

    # Define Local Constant
    PATTERN_LDM_PRJ_DIR = ".*_"
    PATTERN_COMPILE_WARNING_FILE = "(.*[cprCPR]|build)_error.txt"

    # Define Local Variable
    ldm_dir = ""
    prj_dir = ""
    ldm_prj_dir = ""
    ldm_prj_log_mkmot_dir = ""
    compile_warning_file = ""

    # compile warningファイル取得
    print("\n-----Start get compile warning file-----\n")
    ldm_dir = dest_dir + base_target_ldm_file + "/ldm/"
    if os.path.isdir(ldm_dir) is True:
        prj_dir = re.sub(PATTERN_LDM_PRJ_DIR, "", base_target_ldm_file)
        ldm_prj_dir = ldm_dir + prj_dir
        ldm_prj_log_mkmot_dir = ldm_prj_dir + "/log/mkmot/"
        if os.path.isdir(ldm_prj_log_mkmot_dir) is True:
            for current_dir, sub_dirs, files_list in os.walk(ldm_prj_log_mkmot_dir):
                for file_name in files_list:
                    if re.search(PATTERN_COMPILE_WARNING_FILE, file_name):
                        compile_warning_file = (current_dir + "/" + file_name)
                        print("Target compile warning file:", file_name)

    if compile_warning_file == "":
        compile_warning_status = ERROR

    print("\n-----End get compile warning file-----\n")

    return(compile_warning_file)


def get_rst_file():
    """Get rst file."""
    # Refer Global Variable
    global dest_dir
    global base_target_ldm_file
    global memory_usage_status

    # Define Local Constant
    PATTERN_LDM_PRJ_DIR = ".*_"
    PATTERN_RST_FILE = r".*[cprCPR]\.rst"

    # Define Local Variable
    ldm_dir = ""
    prj_dir = ""
    ldm_prj_dir = ""
    rst_file = ""

    # rstファイル取得
    print("\n-----Start get rst file-----\n")

    ldm_dir = dest_dir + base_target_ldm_file + "/ldm/"
    if os.path.isdir(ldm_dir) is True:
        prj_dir = re.sub(PATTERN_LDM_PRJ_DIR, "", base_target_ldm_file)
        ldm_prj_dir = ldm_dir + prj_dir
        if os.path.isdir(ldm_prj_dir) is True:
            for current_dir, sub_dirs, files_list in os.walk(ldm_prj_dir):
                for file_name in files_list:
                    if re.search(PATTERN_RST_FILE, file_name):
                        rst_file = (current_dir + "/" + file_name)
                        print("Target rst file:", file_name)

    if rst_file == "":
        memory_usage_status = ERROR

    print("\n-----End get rst file-----\n")

    return(rst_file)


def get_compile_warning(compile_warning_file):
    """Get compile warning."""
    # Refer Global Variable
    global compile_warning_status

    # Define Local Constant
    PATTERN_COMPILE_WARNING = r".warning\s#"

    # Define Local Variable
    compile_warning_count = 0
    compile_warning_info_line = ""

    # M-09 compiler warning差分取得
    # compile warning（_error.txt）ファイルの「warning #」の文字列の件数取得
    print("\n-----Start get compile warning-----\n")

    try:
        with open(compile_warning_file, "r", encoding="cp932") as compile_warning_info:
            while True:
                compile_warning_info_line = compile_warning_info.readline()
                if compile_warning_info_line == "":
                    break
                if re.search(PATTERN_COMPILE_WARNING, compile_warning_info_line):
                    compile_warning_count += 1
            print("compile warning:", compile_warning_count, "件")
    except FileNotFoundError as file_open_error:
        compile_warning_status = ERROR
        print(file_open_error, compile_warning_file)

    print("\n-----End get compile warning-----\n")

    return(compile_warning_count)


def get_memory_usage(rst_file):
    """Get memory usage."""
    # Refer Global Variable
    global memory_usage_status

    # Define Local Constant
    PATTERN_ROM_INFO = r"^ROM\s+free\s*.\s*\d+\sbyte\s.\s(\d+.\d+)\s%.\s.\s.used\s*.\s*\d+\sbyte\s.\s(\d+.\d+)\s%.\s.\s.size\s*.\s*\d+\sbyte"
    PATTERN_RAM_INFO = r"^RAM\s+free\s*.\s*\d+\sbyte\s.\s(\d+.\d+)\s%.\s.\s.used\s*.\s*\d+\sbyte\s.\s(\d+.\d+)\s%.\s.\s.size\s*.\s*\d+\sbyte"
    PATTERN_USED_BYTE = r"used\s*.\s*\d+\sbyte"
    PATTERN_USAGE = r"\d+"

    # Define Local Variable
    rst_info_line = ""
    # res_rom_info
    # res_ram_info
    rom_info = ""
    ram_info = ""
    # res_rom_used_byte
    # res_ram_used_byte
    rom_used_byte = ""
    ram_used_byte = ""
    # res_rom_usage
    # res_ram_usage
    memory_usage = {"ROM": 0, "RAM": 0}

    # メモリ使用量（ROM/RAM）抽出
    # rstファイルの差分比較を行いメモリ使用量（ROM/RAM）の差分を取得
    print("\n-----Start get memory(ROM/RAM) used size-----\n")

    try:
        with open(rst_file, "r", encoding="cp932") as rst_info:
            while True:
                rst_info_line = rst_info.readline()
                if rst_info_line == "":
                    break
                if re.search(PATTERN_ROM_INFO, rst_info_line):
                    res_rom_info = re.search(PATTERN_ROM_INFO, rst_info_line)
                    rom_info = res_rom_info.group()
                    res_rom_used_byte = re.search(PATTERN_USED_BYTE, rom_info)
                    rom_used_byte = res_rom_used_byte.group()
                    res_rom_usage = re.search(PATTERN_USAGE, rom_used_byte)
                    memory_usage["ROM"] = int(res_rom_usage.group())
                    print("memory(ROM) used size:", memory_usage["ROM"], "byte")
                elif re.search(PATTERN_RAM_INFO, rst_info_line):
                    res_ram_info = re.search(PATTERN_RAM_INFO, rst_info_line)
                    ram_info = res_ram_info.group()
                    res_ram_used_byte = re.search(PATTERN_USED_BYTE, ram_info)
                    ram_used_byte = res_ram_used_byte.group()
                    res_ram_usage = re.search(PATTERN_USAGE, ram_used_byte)
                    memory_usage["RAM"] = int(res_ram_usage.group())
                    print("memory(RAM) used size:", memory_usage["RAM"], "byte")
    except FileNotFoundError as file_open_error:
        memory_usage_status = ERROR
        print(file_open_error, rst_file)

    print("\n-----End get memory(ROM/RAM) used size-----\n")

    return(memory_usage)


def get_diff_build_log_info():
    """Get diff build log information."""
    # Refer Global Variable
    global after_compile_warning
    global before_compile_warning
    global result_compile_warning
    global after_memory_usage
    global before_memory_usage
    global result_memory_usage
    global compile_warning_status
    global memory_usage_status
    global after_compile_warning_status
    global before_compile_warning_status
    global after_memory_usage_status
    global before_memory_usage_status

    # Define Local Constant

    # Define Local Variable

    # build log差分情報取得処理開始
    print("\n-----Start get diff build log info-----\n")

    # compile warning増加件数差分取得
    if (after_compile_warning_status == NORMAL) and (before_compile_warning_status == NORMAL):
        result_compile_warning = get_diff_compile_warning()

    # memory(ROM/RAM)使用量差分取得
    if (after_memory_usage_status == NORMAL) and (before_memory_usage_status == NORMAL):
        result_memory_usage = get_diff_memory_usage()

    # build log差分情報取得処理終了
    print("\n-----End get diff build log info-----\n")


def get_diff_compile_warning():
    """Get diff compile warning."""
    # Refer Global Variable
    global after_compile_warning
    global before_compile_warning

    # Define Local Constant

    # Define Local Variable
    diff_compile_warning = 0

    # compile warning増加件数差分取得処理開始
    print("\n-----Start get diff compile warning-----\n")

    diff_compile_warning = after_compile_warning - before_compile_warning
    if diff_compile_warning < 0:
        diff_compile_warning = 0
    print("compile warning(diff):+", diff_compile_warning, "件")

    # compile warning増加件数差分取得処理終了
    print("\n-----End get diff compile warning-----\n")

    return(diff_compile_warning)


def get_diff_memory_usage():
    """Get diff memory usage."""
    # Refer Global Variable
    global after_memory_usage
    global before_memory_usage

    # Define Local Constant

    # Define Local Variable
    diff_memory_usage = {"ROM": 0, "RAM": 0}

    # memory(ROM/ARM)使用量差分取得処理開始
    print("\n-----Start get diff memory usage-----\n")

    # memory(ROM/RAM)使用量数差分
    diff_memory_usage["ROM"] = after_memory_usage["ROM"] - before_memory_usage["ROM"]
    diff_memory_usage["RAM"] = after_memory_usage["RAM"] - before_memory_usage["RAM"]
    print("memory(ROM) usage(diff):", diff_memory_usage["ROM"], "byte")
    print("memory(RAM) usage(diff):", diff_memory_usage["RAM"], "byte")

    # memory(ROM/ARM)使用量差分取得処理開始
    print("\n-----End get diff memory usage-----\n")

    return(diff_memory_usage)


def get_sca_info():
    """Get SCA information."""
    # Refer Global Variable
    global smr_dir
    global sca_dir

    # Define Local Constant

    # Define Local Variable

    # SCAログファイル保存フォルダ作成
    sca_dir = smr_dir + "M-09/SCA/"
    if os.path.isdir(sca_dir) is False:
        os.makedirs(sca_dir)

    # CodeSonar警告増加件数取得
    get_codesonar_info()

    # QAC警告増加件数取得
    get_qac_info()

    # CRC警告増加件数取得
    get_crc_info()


def get_codesonar_info():
    """Get CodeSonar information."""
    # Refer Global Variable
    global codesonar_project_csv_file
    global after_svn_rev
    global before_svn_rev
    global result_codesonar_warning
    global codesonar_warning_status
    global after_codesonar_warning_status
    global before_codesonar_warning_status

    # Define Local Constant

    # Define Local Variable
    after_rev_codesonar_warning = 0
    before_rev_codesonar_warning = 0

    # CodeSonar警告増加件数取得処理開始
    print("\n-----Start get CodeSonar info-----\n")

    # CodeSonar警告件数取得（after）
    after_rev_codesonar_warning = get_codesonar_warning(after_svn_rev)
    after_codesonar_warning_status = codesonar_warning_status

    # CodeSonar警告件数取得（before）
    before_rev_codesonar_warning = get_codesonar_warning(before_svn_rev)
    before_codesonar_warning_status = codesonar_warning_status

    if (after_codesonar_warning_status == NORMAL) and (before_codesonar_warning_status == NORMAL):
        result_codesonar_warning = after_rev_codesonar_warning - before_rev_codesonar_warning
        if result_codesonar_warning < 0:
            result_codesonar_warning = 0
        print("CodeSonar warning(Score >= 27):+", result_codesonar_warning, "件")

    # CodeSonar警告増加件数取得処理終了
    print("\n-----End get CodeSonar info-----\n")


def get_codesonar_warning(svn_rev):
    """Get CodeSonar warning."""
    # Refer Global Variable
    global codesonar_project_csv_file
    global sca_dir
    global codesonar_dir
    global codesonar_warning_status

    # Define Local Constant
    PATTERN_SVN_REV = r"^\d+"
    PATTERN_TARGET_SVN_REV_CSV_FILE = "analysis/.*csv"
    PATTERN_CODESONAR_ANALYSIS_SCORE = r"^\d+"
    CODESONAR_ANALYSIS_SCORE_WARNING_THRESHOLD = 27

    # header:analysis,state,started,lines with code,url
    class CODESONAR_ANALYSIS_LIST(enum.Enum):
        ANALYSIS = 0           # 0
        STATE = enum.auto()    # 1
        STARTED = enum.auto()  # 2
        LOC = enum.auto()      # 3
        URL = enum.auto()      # 4

    # header:score,id,class,significance,file,line number,procedure,priority,state,finding,owner,url
    class CODESONAR_ANALYSIS_RESULT(enum.Enum):
        SCORE = 0                   # 0
        ID = enum.auto()            # 1
        CLASS = enum.auto()         # 2
        SIGNIFICANCE = enum.auto()  # 3
        FILE = enum.auto()          # 4
        LOC = enum.auto()           # 5
        PROCEDURE = enum.auto()     # 6
        PRIORITY = enum.auto()      # 7
        STATE = enum.auto()         # 8
        FINDING = enum.auto()       # 9
        OWNER = enum.auto()         # 10
        URL = enum.auto()           # 11

    # Define Local Variable
    codesonar_analysis_project_csv_file = ""
    codesonar_analysis_target_svn_rev_csv_file = ""
    # project_csv_file_data
    # target_svn_rev_csv_file_data
    # reader
    header = []
    project_csv_data_line = []
    # res_svn_rev_no
    svn_rev_no = ""
    target_svn_rev_csv_data_line = []
    # res_target_svn_rev_analysis
    target_svn_rev_analysis = ""
    target_svn_rev_analysis_csv_file = ""
    # res_codesonar_analysis_score
    codesonar_analysis_score = 0
    codesonar_warning_count = 0

    # CodeSonar警告件数取得処理開始
    print("\n-----Start get CodeSonar warning-----\n")

    # CodeSonarログファイル保存フォルダ作成
    codesonar_dir = sca_dir + "codesonar/" + svn_rev + "/"
    if os.path.isdir(codesonar_dir) is False:
        os.makedirs(codesonar_dir)

    codesonar_analysis_project_csv_file = codesonar_dir + "codesonar_project.csv"
    codesonar_analysis_target_svn_rev_csv_file = codesonar_dir + "codesonar_" + svn_rev + ".csv"

    # CodeSonarサーバ上の対象プロジェクトの解析一覧リストCSVファイルをローカルのtmpファイルにコピー
    project_csv_file_data = urllib.request.urlopen(codesonar_project_csv_file).read()
    with open(codesonar_analysis_project_csv_file, "wb") as write_project_csv_file:
        write_project_csv_file.write(project_csv_file_data)

    # CodeSonar対象プロジェクトの解析一覧リストCSVファイルのデータ読み出し（1行単位）
    # header:analysis,state,started,lines with code,url
    # analysis(SVN rev) / url(CSV file path)を参照
    try:
        with open(codesonar_analysis_project_csv_file, "r", encoding="shift_jis", errors="", newline="") as read_project_csv_file:
            reader = csv.reader(read_project_csv_file, delimiter=",", doublequote=True, lineterminator="\r\n", quotechar='"', skipinitialspace=True)
            header = next(reader)
            for project_csv_data_line in reader:
                if re.match(PATTERN_SVN_REV, project_csv_data_line[CODESONAR_ANALYSIS_LIST.ANALYSIS.value]):
                    res_svn_rev_no = re.match(PATTERN_SVN_REV, project_csv_data_line[CODESONAR_ANALYSIS_LIST.ANALYSIS.value])
                    svn_rev_no = res_svn_rev_no.group()
                    if svn_rev_no == svn_rev:
                        if re.search(PATTERN_TARGET_SVN_REV_CSV_FILE, project_csv_data_line[CODESONAR_ANALYSIS_LIST.URL.value]):
                            res_target_svn_rev_analysis = re.search(PATTERN_TARGET_SVN_REV_CSV_FILE, project_csv_data_line[CODESONAR_ANALYSIS_LIST.URL.value])
                            target_svn_rev_analysis = res_target_svn_rev_analysis.group()
                            break
    except FileNotFoundError as file_open_error:
        codesonar_warning_status = ERROR
        print(file_open_error, codesonar_analysis_project_csv_file)

    if target_svn_rev_analysis != "":
        # 共通URLと↑を結合し対象revのcsvファイルパス取得
        # 新規警告のみを取得する際はURL末尾に「&filter=5&prj_filter=11」が必要だがうまく取得できないため全件出力とする
        target_svn_rev_analysis_csv_file = CODESONAR_SERVER_COMMON_URL + target_svn_rev_analysis

        # CodeSonarサーバ上の対象revのCodeSonar解析結果CSVファイルをローカルのtmpファイルにコピー
        target_svn_rev_csv_file_data = urllib.request.urlopen(target_svn_rev_analysis_csv_file).read()
        with open(codesonar_analysis_target_svn_rev_csv_file, "wb") as write_target_svn_rev_csv_file:
            write_target_svn_rev_csv_file.write(target_svn_rev_csv_file_data)

        # 対象revのCodeSonar解析結果のデータ読み出し（1行単位）
        # 2行目以降の1列目の値が27以上の場合にカウントアップ
        # header:score,id,class,significance,file,line number,procedure,priority,state,finding,owner,url
        # scoreを参照
        try:
            with open(codesonar_analysis_target_svn_rev_csv_file, "r", encoding="shift_jis", errors="", newline="") as read_target_svn_rev_csv_file:
                reader = csv.reader(read_target_svn_rev_csv_file, delimiter=",", doublequote=True, lineterminator="\r\n", quotechar='"', skipinitialspace=True)
                header = next(reader)
                for target_svn_rev_csv_data_line in reader:
                    if re.match(PATTERN_CODESONAR_ANALYSIS_SCORE, target_svn_rev_csv_data_line[CODESONAR_ANALYSIS_RESULT.SCORE.value]):
                        res_codesonar_analysis_score = re.match(PATTERN_CODESONAR_ANALYSIS_SCORE, target_svn_rev_csv_data_line[CODESONAR_ANALYSIS_RESULT.SCORE.value])
                        codesonar_analysis_score = int(res_codesonar_analysis_score.group())
                        if codesonar_analysis_score >= CODESONAR_ANALYSIS_SCORE_WARNING_THRESHOLD:
                            codesonar_warning_count += 1
            print("CodeSonar warning(Score >= 27):", codesonar_warning_count, "件")
        except FileNotFoundError as file_open_error:
            codesonar_warning_status = ERROR
            print(file_open_error, codesonar_analysis_target_svn_rev_csv_file)

    else:
        codesonar_warning_status = ERROR
        print("Failed:get CodeSonar warning")

    # CodeSonar警告件数取得処理終了
    print("\n-----End get CodeSonar warning-----\n")

    return(codesonar_warning_count)


def get_qac_info():
    """Get QAC information."""
    # Refer Global Variable

    # Define Local Constant

    # Define Local Variable

    # QAC警告増加件数取得処理開始
    print("\n-----Start get QAC warning-----\n")

    print("未対応.")

    # QAC警告増加件数取得処理終了
    print("\n-----End get QAC warning-----\n")


def get_crc_info():
    """Get CRC information."""
    # Refer Global Variable

    # Define Local Constant

    # Define Local Variable

    # CRC警告増加件数取得処理開始
    print("\n-----Start get CRC warning-----\n")

    print("未対応.")

    # CRC警告増加件数取得処理終了
    print("\n-----End get CRC warning-----\n")


def post2excel():
    """Post Excel file."""
    # Refer Global Variable
    global excel_file_name
    global title
    global member_name
    global local_base_rev
    global target_svn_rev
    global loc_info
    global result_compile_warning
    global result_memory_usage
    global result_codesonar_warning
    global loc_diff_status
    global after_compile_warning_status
    global before_compile_warning_status
    global after_codesonar_warning_status
    global before_codesonar_warning_status
    global after_memory_usage_status
    global before_memory_usage_status

    # Define Local Constant
    # Excel転記対象セル（行列）の初期設定
    ROW_INIT = 6

    class COLUMN(enum.Enum):
        BLANK = enum.auto()                      # 1
        TITLE = enum.auto()                      # 2
        MEMBER_NAME = enum.auto()                # 3
        LOCAL_MODIFY_SVN_REVISION = enum.auto()  # 4
        MODIFY_SVN_REVISION = enum.auto()        # 5
        LOC_DELETE = enum.auto()                 # 6
        LOC_INSERT = enum.auto()                 # 7
        LOC_MODIFY = enum.auto()                 # 8
        QAC_WARNING_DIFF = enum.auto()           # 9
        CRC_WARNING_DIFF = enum.auto()           # 10
        COMPILE_WARNING_DIFF = enum.auto()       # 11
        CODESONAR_WARNING_DIFF = enum.auto()     # 12
        MEMORY_ROM_USAGE_DIFF = enum.auto()      # 13
        MEMORY_ROM_USAGE_BEFORE = enum.auto()    # 14
        MEMORY_ROM_USAGE_AFTER = enum.auto()     # 15
        MEMORY_RAM_USAGE_DIFF = enum.auto()      # 16
        MEMORY_RAM_USAGE_BEFORE = enum.auto()    # 17
        MEMORY_RAM_USAGE_AFTER = enum.auto()     # 18

    # Define Local Variable
    # wb
    # ws
    row_start = ROW_INIT

    # Excel転記処理開始
    print("\n-----Start post SCA information to Excel file-----\n")

    print("Open Excel file.")
    # Excelファイルを開く
    wb = openpyxl.load_workbook(excel_file_name)
    # sheetを読み込む
    ws = wb["M08-M09実績"]

    print("Post SCA information.")
    # Excelの対象箇所（B列6行目以降で空白のセル）を探す
    while (ws.cell(row=row_start, column=COLUMN.TITLE.value).value) is not None:
        row_start += 1

    # 案件名 転記
    ws.cell(row=row_start, column=COLUMN.TITLE.value).value = title

    # 担当者名 転記
    ws.cell(row=row_start, column=COLUMN.MEMBER_NAME.value).value = member_name

    # Local base rev 転記
    ws.cell(row=row_start, column=COLUMN.LOCAL_MODIFY_SVN_REVISION.value).value = local_base_rev

    # SVN UP rev 転記
    ws.cell(row=row_start, column=COLUMN.MODIFY_SVN_REVISION.value).value = target_svn_rev

    # LOC（コード差分）情報転記
    if loc_diff_status == NORMAL:
        # delete
        ws.cell(row=row_start, column=COLUMN.LOC_DELETE.value).value = loc_info["delete"]
        # insert
        ws.cell(row=row_start, column=COLUMN.LOC_INSERT.value).value = loc_info["insert"]
        # modify
        ws.cell(row=row_start, column=COLUMN.LOC_MODIFY.value).value = loc_info["modify"]

    # SCA情報転記
    # QAC増加件数転記
    # ws.cell(row=row_start, column=COLUMN.QAC_WARNING_DIFF.value).value = int(QAC差分件数)
    # CRC増加件数転記
    # ws.cell(row=row_start, column=COLUMN.CRC_WARNING_DIFF.value).value = int(CRC差分件数)

    # compiler warning増加件数転記
    if (after_compile_warning_status == NORMAL) and (before_compile_warning_status == NORMAL):
        ws.cell(row=row_start, column=COLUMN.COMPILE_WARNING_DIFF.value).value = result_compile_warning

    if (after_codesonar_warning_status == NORMAL) and (before_codesonar_warning_status == NORMAL):
        # CodeSonar増加件数転記
        ws.cell(row=row_start, column=COLUMN.CODESONAR_WARNING_DIFF.value).value = result_codesonar_warning

    # MEMORY(ROM/RAM)情報転記
    if after_memory_usage_status == NORMAL:
        # MEMORY(ROM）使用量（after）転記
        ws.cell(row=row_start, column=COLUMN.MEMORY_ROM_USAGE_AFTER.value).value = after_memory_usage["ROM"]
        # MEMORY(RAM）使用量（after）転記
        ws.cell(row=row_start, column=COLUMN.MEMORY_RAM_USAGE_AFTER.value).value = after_memory_usage["RAM"]

    if before_memory_usage_status == NORMAL:
        # MEMORY(ROM）使用量（before）転記
        ws.cell(row=row_start, column=COLUMN.MEMORY_ROM_USAGE_BEFORE.value).value = before_memory_usage["ROM"]
        # MEMORY(RAM）使用量（before）転記
        ws.cell(row=row_start, column=COLUMN.MEMORY_RAM_USAGE_BEFORE.value).value = before_memory_usage["RAM"]

    # if (after_memory_usage_status == NORMAL) and (before_memory_usage_status == NORMAL):
        # MEMORY(ROM）使用量差分転記
        # ws.cell(row=row_start, column=COLUMN.MEMORY_ROM_USAGE_DIFF.value).value = result_memory_usage["ROM"]
        # MEMORY(RAM）使用量差分転記
        # ws.cell(row=row_start, column=COLUMN.MEMORY_RAM_USAGE_DIFF.value).value = result_memory_usage["RAM"]

    print("Save excel file.")

    # ファイルを保存する
    wb.save(filename=excel_file_name)

    # ファイルを閉じる
    wb.close()

    # Excel転記処理終了
    print("\n-----End post SCA information to Excel file-----\n")


if __name__ == "__main__":
    main()
