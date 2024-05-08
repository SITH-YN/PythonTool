# -*- coding: utf-8 -*-
"""template."""

# Standard Library
import csv
# import datetime
import enum
# import glob
# import os
import re
# import shutil
import sys
# import time
import tkinter

# 3rd Party Library
import openpyxl


# Global Constant Define
OK = 0
NG = 1
NORMAL = 0
ERROR = 1


# Global Variable


# Define Function
def main():
    """Call Main function."""
    # Refer Global Variable

    # Define Local Constant
    ROOT_DIR = "C:/"

    # Define Local Variable
    work_dir = ""
    excel_file_typ = []
    excel_file_dir = ""
    excel_file_path = ""
    memmap_setting_list = []
    memmap_list = []

    print("\n-----Start Process-----\n")

    print("Please Input Work Dir.")
    print("Example:C:/GitHub/SITH-YN/PythonTool/MemoryMapGenerate/work")
    work_dir = tkinter.filedialog.askdirectory(initialdir=ROOT_DIR)
    # work_dir = input()
    # Debug用ファイルパス設定
    # work_dir = "C:/GitHub/SITH-YN/PythonTool/MemoryMapGenerate/work"
    # print("Work Dir:", work_dir)
    # if (os.path.exists(work_dir) is False):
    #     os.makedirs(work_dir)

    # Open Part Number Excel File
    print("Please Input Memory Map Excel File.")
    excel_file_typ = [("Excel File", "*.xlsx;*.xlsm;*.xls")]
    excel_file_dir = work_dir
    # Debug用ファイルパス設定
    # excel_file_path = excel_file_dir + "/メモリマップ生成ツール_イメージ.xlsm"
    # try:
    excel_file_path = tkinter.filedialog.askopenfilename(filetypes=excel_file_typ, initialdir=excel_file_dir)
    # except Exception as e:
    #     print("存在しないまたはアクセスできないディレクトリを指定しています。")
    #     print(e)
    #     sys.exit(1)

    print("Set Excel File Path.")
    print(excel_file_path)

    # Get Memory Map Setting Information
    print("Get Memory Map Setting Information From Excel File.")
    memmap_setting_list = get_memmap_setting(excel_file_path)

    # Get Memory Map Information
    print("Get Memory Map Information From Excel File.")
    memmap_list = get_memmap_info(excel_file_path, memmap_setting_list)

    # Output ld File
    print("Output Memory Map Information ld File.")
    output_ld(memmap_setting_list, memmap_list)

    print("\n-----End Process-----\n")


def get_memmap_setting(excel_file_path):
    """Get Memory Map Information From Excel File."""
    # Refer Global Variable

    # Define Local Constant
    PATTERN_PART_NUMBER_DATA = r"[^a-fA-F\d]"
    ROW_INIT = 1                               # Default Row
    ROW_MEMMAP_INFO = 6                        # Memory Map Information Row

    class ExcelColumn(enum.Enum):
        """Column Of Memory Map Information Matrix."""

        AREA_ADDRESS = enum.auto()                   # 1(A列)
        AREA = enum.auto()                           # 2(B列)
        LD_FILE_ID = enum.auto()                     # 3(C列)
        MEMBLOCK = enum.auto()                       # 4(D列)
        MEMBLOCK_PREFIX_START_ADDRESS = enum.auto()  # 5(E列)
        MEMBLOCK_START_ADDRESS = enum.auto()         # 6(F列)
        MEMBLOCK_PREFIX_LENGTH = enum.auto()         # 7(G列)
        MEMBLOCK_LENGTH = enum.auto()                # 8(H列)
        PREFIX_MEMORY_CATEGORY = enum.auto()         # 9(I列)
        MEMORY_CATEGORY = enum.auto()                # 10(J列)
        SUFFIX_MEMORY_CATEGORY = enum.auto()         # 11(K列)
        SECTION = enum.auto()                        # 12(L列)
        SECTION_START_EXPRESSION = enum.auto()       # 13(M列)
        SECTION_ATTRIBUTES = enum.auto()             # 14(N列)
        SECTION_COLON = enum.auto()                  # 15(O列)
        SECTION_CONTENTS = enum.auto()               # 16(P列)
        SECTION_LAYOUT = enum.auto()                 # 17(Q列)
        PREFIX_SECTION_CATEGORY = enum.auto()        # 18(R列)
        SECTION_CATEGORY = enum.auto()               # 19(S列)
        SUFFIX_SECTION_CATEGORY = enum.auto()        # 20(T列)
        MEMORY_CALC_EXCLUDE = enum.auto()            # 21(U列)
        PREFIX_COMMENT = enum.auto()                 # 22(V列)
        COMMENT = enum.auto()                        # 23(W列)
        SUFFIX_COMMENT = enum.auto()                 # 24(X列)

    # Define Local Variable
    sheet_name = ""
    row_index = ROW_INIT
    target_column_list = []
    target_column_num = 0
    project_name_list = []
    compile_type_list = []
    id = ""
    part_number_name = ""
    part_number_data = ""
    tmp_part_number_data = ""
    part_number_data_length = ""
    part_number_data_format = ""
    part_number_target = ""
    header_list = ["Project", "Compile Type"]
    part_number_name_list = []
    part_number_data_length_list = []
    part_number_data_format_list = []
    part_number_target_list = []
    part_number_data_list = []
    part_number_name_2d_list = []
    part_number_data_length_2d_list = []
    part_number_data_format_2d_list = []
    part_number_target_2d_list = []
    part_number_data_2d_list = []
    all_part_number_data_list = []
    all_part_number_data_2d_list = []
    ld_file_path = ""

    print("\n-----Start Get Memory Map Information-----\n")

    # Open Excel File
    wb_memmap = openpyxl.load_workbook(excel_file_path, data_only=True)

    sheet_name = "メモリマップ定義"
    ws_wb_memmap = wb_memmap[sheet_name]

    # Get Target Column
    row_index = ROW_MEMMAP_INFO
    target_column_list = []
    target_column_num = 0
    project_name_list = []
    compile_type_list = []

    # Get Part Number Information
    for index in range(target_column_num):
        row_index = ROW_MEMMAP_INFO
        part_number_name_list = []
        part_number_data_format_list = []
        part_number_data_length_list = []
        part_number_target_list = []
        part_number_data_list = []
        all_part_number_data_list = []

        while (ws_wb_memmap.cell(row=row_index, column=ExcelColumn.TARGET.value).value is not None):
            if (ws_wb_memmap.cell(row=row_index, column=ExcelColumn.TARGET.value).value == "○"):
                if (index < 1):
                    id = ws_wb_memmap.cell(row=row_index, column=ExcelColumn.ID.value).value
                    header_list.append(id)
                part_number_name = ws_wb_memmap.cell(row=row_index, column=ExcelColumn.PART_NUMBER_NAME.value).value
                part_number_data_format = ws_wb_memmap.cell(row=row_index, column=ExcelColumn.DATA_FORMAT.value).value
                part_number_data_length = ws_wb_memmap.cell(row=row_index, column=ExcelColumn.DATA_LENGTH.value).value
                part_number_target = ws_wb_memmap.cell(row=row_index, column=ExcelColumn.TARGET.value).value
                part_number_data = ws_wb_memmap.cell(row=row_index, column=target_column_list[index]).value
                tmp_part_number_data = re.sub("0x", "", part_number_data)
                part_number_data = re.sub(PATTERN_PART_NUMBER_DATA, "", tmp_part_number_data)

                part_number_name_list.append(part_number_name)
                part_number_data_format_list.append(part_number_data_format)
                part_number_data_length_list.append(part_number_data_length)
                part_number_target_list.append(part_number_target)
                part_number_data_list.append(part_number_data)

            row_index += 1

        part_number_name_2d_list.insert(index, part_number_name_list)
        part_number_data_format_2d_list.insert(index, part_number_data_format_list)
        part_number_data_length_2d_list.insert(index, part_number_data_length_list)
        part_number_target_2d_list.insert(index, part_number_target_list)
        part_number_data_2d_list.insert(index, part_number_data_list)

        all_part_number_data_list.append(project_name_list[index])
        all_part_number_data_list += compile_type_list[index]
        all_part_number_data_list += part_number_data_list
        all_part_number_data_2d_list.append(all_part_number_data_list)

    # ldファイルの見出しヘッダ情報挿入
    all_part_number_data_2d_list.insert(0, header_list)

    print("Memory Map Information\n")
    print(all_part_number_data_2d_list)

    # Close Excel File
    wb_memmap.close()

    print("\n-----End Get Memory Map Information-----\n")

    return (all_part_number_data_2d_list)


def get_memmap_info(excel_file_path):
    """Get Memory Map Information From Excel File."""
    # Refer Global Variable

    # Define Local Constant
    PATTERN_PART_NUMBER_DATA = r"[^a-fA-F\d]"
    ROW_INIT = 1                               # Default Row
    ROW_MEMMAP_INFO = 6                        # Memory Map Information Row

    class ExcelColumn(enum.Enum):
        """Column Of Memory Map Information Matrix."""

        AREA_ADDRESS = enum.auto()                   # 1(A列)
        AREA = enum.auto()                           # 2(B列)
        LD_FILE_ID = enum.auto()                     # 3(C列)
        MEMBLOCK = enum.auto()                       # 4(D列)
        MEMBLOCK_PREFIX_START_ADDRESS = enum.auto()  # 5(E列)
        MEMBLOCK_START_ADDRESS = enum.auto()         # 6(F列)
        MEMBLOCK_PREFIX_LENGTH = enum.auto()         # 7(G列)
        MEMBLOCK_LENGTH = enum.auto()                # 8(H列)
        PREFIX_MEMORY_CATEGORY = enum.auto()         # 9(I列)
        MEMORY_CATEGORY = enum.auto()                # 10(J列)
        SUFFIX_MEMORY_CATEGORY = enum.auto()         # 11(K列)
        SECTION = enum.auto()                        # 12(L列)
        SECTION_START_EXPRESSION = enum.auto()       # 13(M列)
        SECTION_ATTRIBUTES = enum.auto()             # 14(N列)
        SECTION_COLON = enum.auto()                  # 15(O列)
        SECTION_CONTENTS = enum.auto()               # 16(P列)
        SECTION_LAYOUT = enum.auto()                 # 17(Q列)
        PREFIX_SECTION_CATEGORY = enum.auto()        # 18(R列)
        SECTION_CATEGORY = enum.auto()               # 19(S列)
        SUFFIX_SECTION_CATEGORY = enum.auto()        # 20(T列)
        MEMORY_CALC_EXCLUDE = enum.auto()            # 21(U列)
        PREFIX_COMMENT = enum.auto()                 # 22(V列)
        COMMENT = enum.auto()                        # 23(W列)
        SUFFIX_COMMENT = enum.auto()                 # 24(X列)

    # Define Local Variable
    sheet_name = ""
    row_index = ROW_INIT
    target_column_list = []
    target_column_num = 0
    project_name_list = []
    compile_type_list = []
    id = ""
    part_number_name = ""
    part_number_data = ""
    tmp_part_number_data = ""
    part_number_data_length = ""
    part_number_data_format = ""
    part_number_target = ""
    header_list = ["Project", "Compile Type"]
    part_number_name_list = []
    part_number_data_length_list = []
    part_number_data_format_list = []
    part_number_target_list = []
    part_number_data_list = []
    part_number_name_2d_list = []
    part_number_data_length_2d_list = []
    part_number_data_format_2d_list = []
    part_number_target_2d_list = []
    part_number_data_2d_list = []
    all_part_number_data_list = []
    all_part_number_data_2d_list = []

    print("\n-----Start Get Memory Map Information-----\n")

    # Open Excel File
    wb_memmap = openpyxl.load_workbook(excel_file_path, data_only=True)

    sheet_name = "メモリマップ定義"
    ws_wb_memmap = wb_memmap[sheet_name]

    # Get Target Column
    row_index = ROW_MEMMAP_INFO
    target_column_list = []
    target_column_num = 0
    project_name_list = []
    compile_type_list = []

    # Get Part Number Information
    for index in range(target_column_num):
        row_index = ROW_MEMMAP_INFO
        part_number_name_list = []
        part_number_data_format_list = []
        part_number_data_length_list = []
        part_number_target_list = []
        part_number_data_list = []
        all_part_number_data_list = []

        while (ws_wb_memmap.cell(row=row_index, column=ExcelColumn.TARGET.value).value is not None):
            if (ws_wb_memmap.cell(row=row_index, column=ExcelColumn.TARGET.value).value == "○"):
                if (index < 1):
                    id = ws_wb_memmap.cell(row=row_index, column=ExcelColumn.ID.value).value
                    header_list.append(id)
                part_number_name = ws_wb_memmap.cell(row=row_index, column=ExcelColumn.PART_NUMBER_NAME.value).value
                part_number_data_format = ws_wb_memmap.cell(row=row_index, column=ExcelColumn.DATA_FORMAT.value).value
                part_number_data_length = ws_wb_memmap.cell(row=row_index, column=ExcelColumn.DATA_LENGTH.value).value
                part_number_target = ws_wb_memmap.cell(row=row_index, column=ExcelColumn.TARGET.value).value
                part_number_data = ws_wb_memmap.cell(row=row_index, column=target_column_list[index]).value
                tmp_part_number_data = re.sub("0x", "", part_number_data)
                part_number_data = re.sub(PATTERN_PART_NUMBER_DATA, "", tmp_part_number_data)

                part_number_name_list.append(part_number_name)
                part_number_data_format_list.append(part_number_data_format)
                part_number_data_length_list.append(part_number_data_length)
                part_number_target_list.append(part_number_target)
                part_number_data_list.append(part_number_data)

            row_index += 1

        part_number_name_2d_list.insert(index, part_number_name_list)
        part_number_data_format_2d_list.insert(index, part_number_data_format_list)
        part_number_data_length_2d_list.insert(index, part_number_data_length_list)
        part_number_target_2d_list.insert(index, part_number_target_list)
        part_number_data_2d_list.insert(index, part_number_data_list)

        all_part_number_data_list.append(project_name_list[index])
        all_part_number_data_list += compile_type_list[index]
        all_part_number_data_list += part_number_data_list
        all_part_number_data_2d_list.append(all_part_number_data_list)

    # ldファイルの見出しヘッダ情報挿入
    all_part_number_data_2d_list.insert(0, header_list)

    print("Memory Map Information\n")
    print(all_part_number_data_2d_list)

    # Close Excel File
    wb_memmap.close()

    print("\n-----End Get Memory Map Information-----\n")

    return (all_part_number_data_2d_list)


def output_ld(ld_file_path, memmap_list):
    """Output Memory Map Information To ld File."""
    # Refer Global Variable

    # Define Local Constant

    # Define Local Variable

    print("\n-----Start Output ld File-----\n")

    try:
        with open(ld_file_path, "w", newline="") as f_object:
            writer_object = csv.writer(f_object)
            writer_object.writerows(memmap_list)
    except Exception as e:
        print("ldファイルを新規に作成できません。")
        print("ldファイルが存在するか確認してください。")
        print(e)
        sys.exit(1)

    print("\n-----End Output ld File-----\n")


if __name__ == "__main__":
    main()
