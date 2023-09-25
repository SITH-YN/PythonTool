"""Support Write E-Mail."""

# Standard Library
import csv
import enum
import os
# import re
import shutil
import subprocess
from tkinter import filedialog

# 3rd partyライブラリ
import openpyxl
import win32com.client


# Define Global Constant
NORMAL = 0
ERROR = 1


# Define Global Variable


# Define Function
def main():
    """Call Main function."""
    # Refer Global Variable

    # Define Local Constant
    # メール本文テンプレート保存ディレクトリ
    MAIL_BODY_DIR = "//asds/電子センター(技術部)/TJ000/06他社授受_D/A_008_NSCS/02_共有情報/0202_業務関連情報/020209_設計補助業務/email_template/"
    # 社内DR妥当性NG
    MAIL_BODY_CASE0_PATH = MAIL_BODY_DIR + "社内DR妥当性NG.txt"
    # 社内DR指摘未刈り取り
    MAIL_BODY_CASE1_PATH = MAIL_BODY_DIR + "社内DR指摘未刈り取り.txt"
    # 課題管理表未確認
    MAIL_BODY_CASE2_PATH = MAIL_BODY_DIR + "課題管理表未確認.txt"
    # リスク管理表未確認
    MAIL_BODY_CASE3_PATH = MAIL_BODY_DIR + "リスク管理表未確認.txt"
    # 週報送付
    MAIL_BODY_CASE4_PATH = MAIL_BODY_DIR + "週報送付.txt"

    # Define Local Variable
    support_type = -1
    to_address = ""
    cc_address = ""
    Attachment_file_path = ""

    # 処理開始
    print("\n-----Start Support Write E-Mail-----\n")

    # フォローメール種別の設定
    print("Please input Support Type[0/1/2/3/4].")
    print("Example:\n 0:社内DR妥当性NG\n 1:社内DR指摘未刈り取り\n 2:課題管理表未確認\n 3:リスク管理表未確認\n 4:週報送付\n")
    support_type = int(input())

    # Outlookのオブジェクト設定
    outlook = win32com.client.Dispatch("Outlook.Application")
    mymail = outlook.CreateItem(0)
    mymail.BodyFormat = 1

    # 署名設定
    # sign = '''
    # 株式会社 ホゲホゲドットコム
    # 世界のナベヒロ
    # '''

    # mymail.To = "yota.nakamura@aisin.co.jp; takeshi.tazaki@aisin.co.jp"
    # mymail.cc = "keiji.kondo@aisin.co.jp"
    # mymail.Bcc = "bar@hoge.com"
    # path = r"C:\\Users\watya\Desktop\hogehoge.txt" # 添付ファイルは絶対パスで指定
    # mymail.Attachments.Add (path)
    # mymail.Body = '''各位
    # お疲れ様です。

    # 以上、よろしくお願いいたします。
    # '''+ "\n" +sign

    match support_type:
        # 社内DR妥当性NG
        case 0:
            to_address = ""
            cc_address = "keiji.kondo@aisin.co.jp; yusaku.fujimoto@aisin.co.jp"
            mymail.To = to_address
            mymail.cc = cc_address
            mymail.Subject = "[NSCS]社内DR妥当性"

            with open(MAIL_BODY_CASE0_PATH, encoding="shift_jis") as textfile:
                mail_body = textfile.read()

            mymail.Body = mail_body

        # 社内DR指摘未刈り取り
        case 1:
            to_address = ""
            cc_address = ""
            mymail.To = to_address
            mymail.cc = cc_address
            mymail.Subject = "[NSCS]社内DR指摘"

            with open(MAIL_BODY_CASE1_PATH, encoding="shift_jis") as textfile:
                mail_body = textfile.read()

            mymail.Body = mail_body

        # 課題管理表未確認
        case 2:
            # SMR一覧チェック
            check_SMR_list()

            to_address = ""
            cc_address = "keiji.kondo@aisin.co.jp; yusaku.fujimoto@aisin.co.jp"
            mymail.To = to_address
            mymail.cc = cc_address
            mymail.Subject = "[NSCS]課題管理表の確認依頼"

            with open(MAIL_BODY_CASE2_PATH, encoding="shift_jis") as textfile:
                mail_body = textfile.read()

            mymail.Body = mail_body

        # リスク管理表未確認
        case 3:
            to_address = ""
            cc_address = "keiji.kondo@aisin.co.jp; yusaku.fujimoto@aisin.co.jp"
            mymail.To = to_address
            mymail.cc = cc_address
            mymail.Subject = "[NSCS]リスク管理表の確認依頼"

            with open(MAIL_BODY_CASE3_PATH, encoding="shift_jis") as textfile:
                mail_body = textfile.read()

            mymail.Body = mail_body

        # 週報送付
        case 4:
            # 週報ファイル圧縮（7z形式で圧縮後7z_形式に拡張子変更）
            compress_file = create_weekly_report()

            to_address = "kondok@nscs.jp"
            cc_address = "keiji.kondo@aisin.co.jp; embsol_sup@nscs.jp; satoshi.nakagawa@aisin.co.jp; taketo.kajihara@aisin.co.jp"
            mymail.To = to_address
            mymail.cc = cc_address
            mymail.Subject = "【社外送信】[NSCS勤怠] 週報送付(yymmdd-yymmdd)"

            with open(MAIL_BODY_CASE4_PATH, encoding="shift_jis") as textfile:
                mail_body = textfile.read()

            mymail.Body = mail_body

            # 添付ファイルは絶対パスで指定
            Attachment_file_path = compress_file
            mymail.Attachments.Add(Attachment_file_path)

        # default
        case _:
            print("Invalid parameter(support_type).")

    # 出来上がったメール確認（True：確認画面立ち上がり後アクションしないと処理が進まない/False：確認画面立ち上がり後処理が進む）
    mymail.Display(False)
    # 確認せず送信する場合は、mymail.Display(True/False)を消して、下記コードを使用する。
    # mymail.Send()

    # フォルダ名変更
    # 圧縮ファイル移動

    # 処理終了
    print("\n-----End Support Write E-Mail-----\n")


def check_SMR_list():
    """Check SMR List Excel file."""
    # Refer Global Variable

    # Define Local Constant
    SMR_LIST_DIR = "//asds/電子センター(技術部)/TJ000/06他社授受_D/A_008_NSCS/02_共有情報/0203_業務/"

    class COLUMN(enum.Enum):
        BLANK = enum.auto()              # 1
        TEAM = enum.auto()               # 2
        WEEK_1ST = enum.auto()           # 3
        WEEK_2ND = enum.auto()           # 4
        WEEK_3RD = enum.auto()           # 5
        WEEK_4TH = enum.auto()           # 6
        WEEK_5TH = enum.auto()           # 7
        PROBLEM_LIST_PATH = enum.auto()  # 8

    # Excel転記対象セル（行列）の初期設定
    ROW_INIT = 1                         # 1行目

    # Define Local Variable
    date = ""
    split_date = ""
    year = ""
    month = ""
    # day = ""
    month_number = ""
    excel_file_typ = ""
    excel_file_dir = ""
    excel_file_name = ""
    smr_list_sheet = ""
    wb_smr_list = ""
    ws_smr_list = ""
    row_start = ROW_INIT
    problem_list_path = ""
    target_team = ""
    problem_list_check_result = ()

    # SMR一覧チェック処理開始
    print("\n-----Start Check SMR List Excel file-----\n")

    # Check Date(Year/Month/Day)指定
    print("Please input year/month/day(yyyy/mm/dd).")
    print("Example: 2023/03/05")
    date = input()
    split_date = date.split("/")
    year = split_date[0]
    month = split_date[1]
    # day = split_date[2]

    if month[0] == "0":
        month_number = month.replace("0", "")
    else:
        month_number = month

    # Excelファイルパス設定
    print("Please input SMR List excel file.")
    excel_file_typ = [("excelファイル", "*.xlsx")]
    excel_file_dir = SMR_LIST_DIR
    excel_file_name = filedialog.askopenfilename(filetypes=excel_file_typ, initialdir=excel_file_dir)
    print("Set SMR List excel file.")
    print(excel_file_name)

    print("Open SMR List Excel file.")
    # Excelファイルを開く
    wb_smr_list = openpyxl.load_workbook(excel_file_name, data_only=True)
    smr_list_sheet = year + "." + month_number
    # sheetを読み込む
    ws_smr_list = wb_smr_list[smr_list_sheet]

    # 課題管理表チェック処理
    # SMR一覧のB列1行目以降で【課題管理】タグ定義セルを探す
    while (ws_smr_list.cell(row=row_start, column=COLUMN.TEAM.value).value != "【課題管理】"):
        row_start += 1
    # 見出しをスキップして実際の開始行を設定
    row_start += 3

    # チェック対象T数分ループ
    # SMR一覧のB列の【課題管理】定義以降空白セルでなければ課題管理表チェック実施
    while (ws_smr_list.cell(row=row_start, column=COLUMN.TEAM.value).value) is not None:
        # 対象Team名取得
        target_team = ws_smr_list.cell(row=row_start, column=COLUMN.TEAM.value).value

        # 対象T行D列の課題管理表パスを取得
        problem_list_path = ws_smr_list.cell(row=row_start, column=COLUMN.PROBLEM_LIST_PATH.value).value

        # 課題管理表チェック処理
        problem_list_check_result = problem_list_check_result + check_problem_list(target_team, problem_list_path)

        row_start += 1

    # チェック結果CSV出力
    print(problem_list_check_result)
    csv_path = SMR_LIST_DIR + "ProblemList_CheckResult.csv"
    with open(csv_path, "w", newline="") as file:
        writer = csv.writer(file)
        writer.writerows(problem_list_check_result)

    print(csv_path)

    # ファイルを保存する
    # print("Save excel file.")
    # wb_smr_list.save(filename=excel_file_name)

    # ファイルを閉じる
    wb_smr_list.close()

    # SMR一覧チェック処理終了
    print("\n-----End Check SMR List Excel file-----\n")


def check_problem_list(target_team, problem_list_path):
    """Check Problem List Excel file."""
    # Refer Global Variable

    # Define Local Constant
    class COLUMN(enum.Enum):
        BLANK = enum.auto()         # 1
        REVIEWER = enum.auto()      # 2
        CHECK_DATE = enum.auto()    # 3
        CHECK_STATUS = enum.auto()  # 4

    # Excel転記対象セル（行列）の初期設定
    ROW_INIT = 1                    # 1行目

    # Define Local Variable
    problem_list_sheet = ""
    wb_problem_list = ""
    ws_problem_list = ""
    row_start = ROW_INIT
    team = ""
    reviewer = ""
    check_status = ""
    check_result = ()

    # 課題管理表チェック処理開始
    print("\n-----Start Check Problem List Excel file-----\n")

    # 対象Tの課題管理表を開く
    print("Open Problem List Excel file.")
    print(problem_list_path)
    # Excelファイルを開く
    wb_problem_list = openpyxl.load_workbook(problem_list_path, data_only=True)
    # 課題管理表は全T共通フォーマットのためシート名固定
    problem_list_sheet = "残項目一覧"
    # sheetを読み込む
    ws_problem_list = wb_problem_list[problem_list_sheet]

    # SMR一覧のB列1行目以降で確認者定義セルを探す
    while (ws_problem_list.cell(row=row_start, column=COLUMN.REVIEWER.value).value != "確認者"):
        row_start += 1
    # 見出しをスキップして実際の開始行を設定
    row_start += 1

    # SMR一覧のB列の確認者定義セル以降空白セルでなければH列のチェックステータス取得
    while (ws_problem_list.cell(row=row_start, column=COLUMN.REVIEWER.value).value) is not None:
        team = target_team
        reviewer = ws_problem_list.cell(row=row_start, column=COLUMN.REVIEWER.value).value
        check_status = ws_problem_list.cell(row=row_start, column=COLUMN.CHECK_STATUS.value).value
        check_result = check_result + ((team, reviewer, check_status), )

        row_start += 1

    # ファイルを閉じる
    wb_problem_list.close()

    # 課題管理表チェック処理終了
    print("\n-----End Check Problem List Excel file-----\n")

    return check_result


def create_weekly_report():
    """Create Weekly Report."""
    # Refer Global Variable

    # Define Local Constant
    WORKING_HOUR_MANAGEMET_DIR = "//asds/電子センター(技術部)/TJ000/06他社授受_D/A_008_NSCS/03_工数/0301_工数管理/週報作成/"
    WEEKLY_REPORT_DIR = "//asds/電子センター(技術部)/TJ000/06他社授受_D/A_008_NSCS/04_週報/"

    # Define Local Variable
    date = ""
    split_date = ""
    year = ""
    month = ""
    day = ""
    month_number = ""
    week = ""
    year_dir = ""
    month_dir = ""
    week_dir = ""
    copy_src_file = ""
    copy_dest_file = ""
    compress_src_file = ""
    compress_dest_file = ""

    # 週報ファイル生成処理開始
    print("\n-----Start Create Weekly Report file-----\n")

    # Create Date(Year/Month/Day)指定
    print("Please input create week folder name -> year/month/day(yyyy/mm/dd).")
    print("Example: 2023/03/05")
    date = input()
    split_date = date.split("/")
    year = split_date[0]
    month = split_date[1]
    day = split_date[2]

    if month[0] == "0":
        month_number = month.replace("0", "")
    else:
        month_number = month

    week = "～" + year + month + day

    year_dir = WEEKLY_REPORT_DIR + year + "/"
    month_dir = year_dir + month_number + "月/"
    week_dir = month_dir + week + "/"

    # 指定したWeekフォルダが既に存在している場合は既存フォルダを削除
    if os.path.isdir(week_dir) is True:
        shutil.rmtree(week_dir)
    # 工数管理\週報作成フォルダ内の週報ファイルをWeekフォルダへコピー
    shutil.copytree(WORKING_HOUR_MANAGEMET_DIR, week_dir)
    # 指定した週報送付フォルダが既に存在している場合は既存フォルダを削除
    # Weekフォルダ内の週報ファイルを週報送付フォルダへコピー（～yyyymmdd→NSCS_週報送付）
    copy_src_file = week_dir
    copy_dest_file = month_dir + "NSCS_週報送付"
    if os.path.isdir(copy_dest_file) is True:
        shutil.rmtree(copy_dest_file)
    shutil.copytree(copy_src_file, copy_dest_file)

    # 7zipで7z形式に圧縮
    compress_src_file = copy_dest_file
    compress_dest_file = compress_src_file + ".7z"
    compress_7zip(compress_src_file, compress_dest_file)
    shutil.rmtree(compress_src_file)
    # 7z_形式にファイル拡張子変更
    copy_src_file = compress_dest_file
    copy_dest_file = copy_src_file + "_"
    # 指定した7z_形式の圧縮ファイルが既に存在している場合は既存の圧縮ファイルを削除
    if os.path.isfile(copy_dest_file) is True:
        os.remove(copy_dest_file)
    os.rename(copy_src_file, copy_dest_file)

    # 週報ファイル生成処理終了
    print("\n-----End Create Weekly Report file-----\n")

    return copy_dest_file


def compress_7zip(src_file, dest_file):
    """7-Zip File compression."""
    # Refer Global Variable

    # Define Local Constant
    TOOLPATH_X86_7ZIP = "C:/Program Files (x86)/7-Zip/7z.exe"
    TOOLPATH_X64_7ZIP = "C:/Program Files/7-Zip/7z.exe"

    # Define Local Variable
    # cmd = ""
    # completed_process
    return_code = NORMAL
    toolpath_7zip = ""

    # ファイル圧縮処理開始
    print("\n-----Start compress file-----\n")

    # 7-zipアプリケーションの設定
    if os.path.isfile(TOOLPATH_X86_7ZIP) is True:
        toolpath_7zip = TOOLPATH_X86_7ZIP
    elif os.path.isfile(TOOLPATH_X64_7ZIP) is True:
        toolpath_7zip = TOOLPATH_X64_7ZIP
    else:
        toolpath_7zip = ""
        print("Not found 7-zip appication.")
        print("Please install 7-zip.")

    if toolpath_7zip != "":
        # 週報フォルダの圧縮
        # ↓スペースのsplitではC:/Program Files部分を認識できないため直接コマンドを指定
        print("Compressing file.")
        # cmd = CMD_DECOMPRESSION + " " + dest_file + " -o " + dest_dir + " -r " + TARGET_FILE_EXTENSION
        # completed_process = subprocess.run(cmd.split(), stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        # completed_process = subprocess.run(toolpath_7zip + " x " + dest_file + " -o" + dest_dir + " -r " + TARGET_FILE_EXTENSION, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        completed_process = subprocess.run(toolpath_7zip + " a " + dest_file + " " + src_file, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        return_code = completed_process.returncode

        if return_code != NORMAL:
            print("Error:Failed file compression.")
        else:
            print("Success:Successed file compression.")

    # 圧縮ファイル解凍処理終了
    print("\n-----End compress file-----\n")


if __name__ == "__main__":
    main()
